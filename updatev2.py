#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Todo 日报回写应用 v6（闭环版）

与 v3 的区别：
- 进展回写（insertProgress）：使用日报 h3 @人名匹配到的 owner.userGuid，而非 Todo 第一个 owner。
- 字段更新（update：promiseDate/riskLevel/title）：使用顺延中实际贡献字段的 owner 身份，不同字段贡献者不同时分次调用。
- promiseDate/riskLevel 按 owner 顺序顺延：第一位 owner 填了用第一位的，没填用第二位的，依次类推。
- 格式错误提醒卡片：发送给 @人名匹配到的 owner.userGuid，而非 Todo 第一个 owner。
- 解析范围：只抓取 h2「平台项目重要任务进展」之后的 h3 @人名区域内的任务。

作用：
1. 根据 config.todo_project_id 先拉取 Todo 项目下的全量任务，建立 id -> Todo 的索引。
2. 根据 mapping_guid 找到各项目当天日报。
3. 从日报 JSON 中解析 TodoList任务进展：h3 @人名标识 owner 区域，🚀 行提取 todo_id、承诺完成时间、风险等级、今日进展。
4. 用解析出的 todo_id 匹配 Todo 索引，用 @人名匹配 owners 列表确定进展回写身份。
5. 匹配成功后：
   - 进展回写使用 @人名匹配到的 owner 的 x-user-guid。
   - 字段更新（promiseDate/riskLevel/title）使用顺延中实际贡献字段的 owner 的 x-user-guid，不同字段贡献者不同时分次调用 update。
6. 若用户填写的承诺完成时间/风险等级与 Todo 当前值一致，则不重复传该字段，并在进展中追加提醒。
7. 未填写的字段保持 None，不传入更新 payload。
8. 若用户填写的承诺完成时间/风险等级格式错误，则通过 @人名匹配到的 owner.userGuid 发送飞书个人卡片提醒。
9. 按 owner 判断“三项是否完整”：进展描述、承诺完成时间、风险等级任一缺失/非法即为个人未更新；任一 owner 未完整更新则任务 title 加【X次未更新】；格式错误加【格式错误】。
10. 进展内容通过 insertProgress 接口写入，promiseDate/riskLevel/title 通过 update 接口更新。
"""

import builtins
import sys
import re
import csv
import json
import time
import traceback
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from collections import OrderedDict

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook

from zenv import get_zdkit_env
from zdbase import ZFile  # 保留平台兼容，不直接依赖

# =============================================================================
# 日志编码兜底
# =============================================================================

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# =============================================================================
# print flush patch
# =============================================================================

if not getattr(builtins.print, "_patched_flush", False):
    _original_print = builtins.print

    def print(*args, **kwargs):
        kwargs.setdefault("flush", True)
        _original_print(*args, **kwargs)

    print._patched_flush = True
    builtins.print = print

# =============================================================================
# 固定 API 路由
# =============================================================================

BASE_URL = "https://workspace.cxmt.com"
TODO_BASE_URL = "http://share.cxmt.com"

TODO_LIST_API = "/openapi/todo/todoList"
TODO_UPDATE_API = "/openapi/todo/update"
TODO_INSERT_PROGRESS_API = "/openapi/todo/insertProgress"

GET_DOC_API = "/platform/ws/noteInfo/getDocJson"
TOKEN_API = "/api/user/platform/getAccessToken"
DOC_TREE_ROUTE = "/platform/api/main/doc/treeList"
SIGNED_URL_ROUTE = "/platform/api/main/storage/getSignedUrl"

MESSAGE_SEND_ROUTE = "/middle/server/api/msg/send"
WORKSPACE_SAVE_ROUTE = "/platform/api/main/workspace/save"
MD_INSERT_ROUTE = "/middle/server/api/file/md/insert"

# =============================================================================
# 固定业务参数
# =============================================================================

TODO_STATUS_LIST = ["Open", "Ongoing", "Delay"]
TODO_IS_LEAF = False
TODO_TYPE = "wbs"

MESSAGE_TEMPLATE_ID = "80"
PLATFORM_TYPE = "all"

RISK_MAP_CN_TO_EN = {
    "高": "high",
    "中": "middle",
    "低": "low",
}

RISK_MAP_EN_TO_CN = {
    "high": "高",
    "middle": "中",
    "low": "低",
}

# 这些内容不应被当作有效进展写回 Todo
DEFAULT_PROGRESS_VALUES = {
    "完成了xxx",
    "开发了xxx",
    "请删除本占位符，并填写今日实际工作进展",
    "请删除本占位符，并填写今日实际完成内容",
    "请在此处填写今日实际工作进展",
    "请在此处填写今日实际完成内容",
}

# =============================================================================
# 全局配置加载
# =============================================================================

zenv_obj = get_zdkit_env()
BASE_URL = zenv_obj.zdkit._http_client.config.get("url")

try:
    with open(config_file.path, "r", encoding="utf-8") as config_fp:
        config = json.load(config_fp)
except Exception as e:
    print(f"❌ 配置文件读取失败: {e}")
    raise

AK = config.get("ak")
SK = config.get("sk")
ORG_GUID = config.get("org_guid")
USER_GUID = config.get("user_guid")

MAPPING_GUID = config.get("mapping_guid")
TARGET_DATE = config.get("target_date") or datetime.now().strftime("%Y-%m-%d")

TODO_PROJECT_ID = config.get("todo_project_id")
if isinstance(TODO_PROJECT_ID, str):
    TODO_PROJECT_ID = [TODO_PROJECT_ID]

# 飞书卡片提醒配置
FORMAT_ERROR_CARD_TITLE = "📍 WBS任务格式填写错误提醒"
FORMAT_ERROR_BUTTON_URL = "http://workshare.cxmt.com/project/list"
FORMAT_ERROR_SENDER_GUID = USER_GUID

if not AK or not SK:
    print("❌ 配置错误：缺少 ak / sk")
    raise ValueError("配置错误：缺少 ak / sk")

if not USER_GUID:
    print("❌ 配置错误：缺少 user_guid")
    raise ValueError("配置错误：缺少 user_guid")

if not MAPPING_GUID:
    print("❌ 配置错误：缺少 mapping_guid")
    raise ValueError("配置错误：缺少 mapping_guid")

if not TODO_PROJECT_ID:
    print("❌ 配置错误：缺少 todo_project_id")
    raise ValueError("配置错误：缺少 todo_project_id")

REPORT_TARGET_PROJECT_GUID = config.get("report_target_project_guid", "")
REPORT_UPDATE_PARENT_GUID = config.get("report_update_parent_guid", "")

REPORT_SENDER_GUID = config.get("report_sender_guid", [])
if isinstance(REPORT_SENDER_GUID, str):
    REPORT_SENDER_GUID = [REPORT_SENDER_GUID]
REPORT_SENDER_GUID = [x for x in REPORT_SENDER_GUID if x]

REPORT_WEBHOOK = config.get("report_webhook", [])
if isinstance(REPORT_WEBHOOK, str):
    REPORT_WEBHOOK = [REPORT_WEBHOOK]
REPORT_WEBHOOK = [x for x in REPORT_WEBHOOK if x]

# =============================================================================
# 通用 HTTP 工具
# =============================================================================

session = requests.Session()

retry_strategy = Retry(
    total=5,
    connect=5,
    read=5,
    backoff_factor=2,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["HEAD", "GET", "POST", "PUT", "DELETE", "OPTIONS"],
    raise_on_status=False,
)

adapter = HTTPAdapter(
    max_retries=retry_strategy,
    pool_connections=10,
    pool_maxsize=10,
)

session.mount("http://", adapter)
session.mount("https://", adapter)


def request_with_retry(method, url, max_retries=3, **kwargs):
    """
    稳定版 HTTP 请求：
    1. 自动 retry
    2. Connection reset 自动恢复
    3. 避免复用坏连接
    """

    kwargs.setdefault("timeout", 30)
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            headers = kwargs.pop("headers", {}) or {}
            headers["Connection"] = "close"
            headers.setdefault("User-Agent", "Mozilla/5.0 PythonRequests")

            if method.lower() == "post":
                response = session.post(url, headers=headers, **kwargs)
            else:
                response = session.get(url, headers=headers, **kwargs)

            return response

        except (
            requests.exceptions.Timeout,
            requests.exceptions.ConnectionError,
            requests.exceptions.ChunkedEncodingError,
        ) as e:
            last_error = e
            wait = min(2 ** attempt, 10)

            print(
                f"⚠️ 请求失败 attempt={attempt}/{max_retries} "
                f"wait={wait}s url={url} error={repr(e)}"
            )

            if attempt < max_retries:
                time.sleep(wait)
            else:
                raise last_error

# =============================================================================
# token / headers
# =============================================================================


def get_access_token():
    print("🔑 获取 AccessToken")

    response = request_with_retry(
        "post",
        BASE_URL + TOKEN_API,
        json={"ak": AK, "sk": SK},
        timeout=30,
    )

    response.raise_for_status()
    response_json = response.json()

    token = (response_json.get("data") or {}).get("accessToken")

    if not token:
        raise Exception(
            f"获取 token 失败: url={BASE_URL + TOKEN_API}, response={response_json}"
        )

    print("✅ AccessToken 获取成功")
    return token


def get_headers(user_guid=None):
    """BASE_URL / Workspace 专用 header。"""
    token = get_access_token()

    return {
        "Access-Token": token,
        "ak": AK,
        "X-User-GUID": user_guid or USER_GUID,
        "Content-Type": "application/json",
    }


def get_todo_headers(user_guid=None):
    """
    TODO_BASE_URL / share 专用 header。

    回写 Todo 时优先使用匹配到的身份 guid。
    拉取 Todo 列表时使用 config.USER_GUID。
    """
    return {
        "x-user-guid": user_guid or USER_GUID,
        "Content-Type": "application/json",
    }


def get_headers_with_ak(user_guid=None):
    """消息发送专用 header，同时携带 Access-Token 和 ak。"""
    token = get_access_token()
    return {
        "Access-Token": token,
        "ak": AK,
        "X-User-GUID": user_guid or USER_GUID,
        "Content-Type": "application/json",
    }


def send_message_api(receiver_guids, title, content, sender_guid="", interactive_content=None):
    """通过平台消息接口发送飞书卡片/文本消息。"""
    payload = {
        "template_id": MESSAGE_TEMPLATE_ID,
        "receiver_guid": receiver_guids,
        "content": content,
        "org_guid": ORG_GUID,
        "title": title,
        "platform_type": PLATFORM_TYPE,
    }
    if interactive_content is not None:
        payload["interactive_content"] = json.dumps(interactive_content, ensure_ascii=False)

    return request_with_retry(
        "post",
        BASE_URL + MESSAGE_SEND_ROUTE,
        headers=get_headers_with_ak(user_guid=sender_guid),
        json=payload,
        timeout=30,
    )


# =============================================================================
# mapping 文件读取
# =============================================================================


def get_signed_url(category_guid):
    print(f"🔗 获取签名 URL: category_guid={category_guid}")

    response = request_with_retry(
        "get",
        BASE_URL + SIGNED_URL_ROUTE,
        headers=get_headers(),
        params={"categoryGuid": category_guid},
        timeout=30,
    )

    response.raise_for_status()
    response_json = response.json()
    signed_url = (response_json.get("data") or {}).get("signedUrl")

    if not signed_url:
        raise Exception(
            f"获取签名 URL 失败: category_guid={category_guid}, response={response_json}"
        )

    return signed_url


def parse_xlsx_mapping(content_bytes):
    workbook = load_workbook(
        filename=BytesIO(content_bytes),
        read_only=True,
        data_only=True,
    )

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValueError("xlsx 文件为空")

    headers = [str(x).strip() if x is not None else "" for x in headers]
    rows = []

    for excel_row in rows_iter:
        row_dict = {}

        for idx, value in enumerate(excel_row):
            if idx >= len(headers):
                continue

            row_dict[headers[idx]] = str(value).strip() if value is not None else ""

        rows.append({
            "dept": row_dict.get("Dept", ""),
            "project_guid": row_dict.get("project_guid", ""),
            "work_log_folder_guid": row_dict.get("work_log_folder_guid", ""),
            "project_name": row_dict.get("project_name", ""),
        })

    return rows


def parse_csv_mapping(text):
    text = text.lstrip("﻿")
    stream = StringIO(text)
    reader = csv.DictReader(stream)

    rows = []

    for row in reader:
        rows.append({
            "dept": row.get("Dept", ""),
            "project_guid": row.get("project_guid", ""),
            "work_log_folder_guid": row.get("work_log_folder_guid", ""),
            "project_name": row.get("project_name", ""),
        })

    return rows


def load_mapping_from_guid(mapping_guid):
    print(f"🚀 开始读取 mapping: {mapping_guid}")

    signed_url = get_signed_url(mapping_guid)

    response = request_with_retry(
        "get",
        signed_url,
        timeout=120,
        stream=True,
        headers={"Connection": "close"},
    )

    response.raise_for_status()
    content = response.content
    signed_url_lower = signed_url.lower()

    if ".xlsx" in signed_url_lower:
        rows = parse_xlsx_mapping(content)
    else:
        rows = parse_csv_mapping(
            content.decode("utf-8-sig", errors="ignore")
        )

    print(f"✅ mapping 读取完成: 有效行数={len(rows)}")
    return rows

# =============================================================================
# Todo API：先按 project_id 拉取全量任务，用于 todo_id 匹配和 header/状态比对
# =============================================================================


def fetch_todo_list():
    print(f"🚀 获取 Todo 列表: projectIds={TODO_PROJECT_ID}")

    request_body = {
        "projectIds": TODO_PROJECT_ID,
        "todoStatus": TODO_STATUS_LIST,
        "isLeaf": TODO_IS_LEAF,
        "type": TODO_TYPE,
    }

    response = request_with_retry(
        "post",
        TODO_BASE_URL + TODO_LIST_API,
        headers=get_todo_headers(),
        json=request_body,
        timeout=60,
    )

    response.raise_for_status()
    response_json = response.json()

    if response_json.get("code") not in (0, 200, "0", "200", None):
        print(
            f"⚠️ Todo API 返回 code 异常: "
            f"code={response_json.get('code')}, msg={response_json.get('msg')}"
        )

    todo_list = response_json.get("data") or []
    print(f"✅ 获取 Todo 数量: {len(todo_list)}")

    return todo_list


def build_todo_index(todo_list):
    """构建 id -> todo 的索引。"""
    todo_index = {}
    duplicate_ids = set()

    for todo in todo_list:
        todo_id = todo.get("id")
        if todo_id in (None, ""):
            continue

        todo_id_int = safe_int(todo_id)
        if todo_id_int is None:
            continue

        if todo_id_int in todo_index:
            duplicate_ids.add(todo_id_int)

        todo_index[todo_id_int] = todo

    if duplicate_ids:
        print(f"⚠️ Todo id 存在重复，已以后出现者覆盖: {sorted(duplicate_ids)}")

    print(f"✅ Todo 索引构建完成: {len(todo_index)} 条")
    return todo_index


def build_dept_user_guid_map(todo_list):
    """
    基于 todo owners 构建：
    dept(fullDeptName) -> [userGuid1, userGuid2...]

    用于：
    根据 mapping.dept 找一个有项目权限的人去读取日报。
    """

    dept_map = OrderedDict()

    for todo in todo_list:
        owners = todo.get("owners", []) or []

        for owner in owners:
            dept = str(owner.get("fullDeptName") or "").strip()
            user_guid = str(owner.get("userGuid") or "").strip()

            if not dept or not user_guid:
                continue

            if dept not in dept_map:
                dept_map[dept] = []

            if user_guid not in dept_map[dept]:
                dept_map[dept].append(user_guid)

    print(f"✅ 部门用户索引构建完成: dept_count={len(dept_map)}")

    return dept_map

def safe_int(value):
    try:
        return int(value)
    except Exception:
        return None


def get_first_owner_user_guid(todo):
    owners = todo.get("owners", []) or []
    for owner in owners:
        user_guid = owner.get("userGuid")
        if user_guid:
            return user_guid
    return None


def get_first_owner_display_name(todo):
    owners = todo.get("owners", []) or []
    for owner in owners:
        return owner.get("displayName") or owner.get("name") or owner.get("loginName") or ""
    return ""


def match_owner_by_name(todo, owner_name):
    """
    根据日报中的 @人名 匹配 todo 的 owners 列表。
    先匹配 owners[].name，未命中则尝试 owners[].englishName。
    返回匹配到的 owner 对象或 None。
    """
    if not owner_name:
        return None
    owners = todo.get("owners", []) or []
    for owner in owners:
        name = str(owner.get("name") or "").strip()
        if name == owner_name:
            return owner
    for owner in owners:
        english_name = str(owner.get("englishName") or "").strip()
        if english_name == owner_name:
            return owner
    return None


def get_todo_code_for_display(todo):
    """用于卡片展示的 Todo code；接口缺失 code 时用 WBS#{id} 兜底。"""
    code = str(todo.get("code") or "").strip()
    if code:
        return code

    todo_id = todo.get("id")
    if todo_id not in (None, ""):
        return f"WBS#{todo_id}"

    return ""


def get_todo_related_project_name(todo, fallback_project_name=""):
    related_project = todo.get("relatedProject") or {}
    return (
        related_project.get("name")
        or fallback_project_name
        or "未命名项目"
    )


def normalize_date_for_compare(date_value):
    text = str(date_value or "").strip()
    if not text:
        return ""

    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"

    m = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})", text)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"

    m = re.match(r"^(\d{4})(\d{2})(\d{2})$", text)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{mo}-{d}"

    return text

# =============================================================================
# 日报查找
# =============================================================================


def get_tree_node_title(node):
    for key in ("dataTitle", "title", "name", "fileName", "filename"):
        value = node.get(key)
        if value:
            return str(value).strip()
    return ""


def get_tree_node_guid(node):
    for key in ("categoryGuid", "dataGuid", "guid", "fileGuid", "id"):
        value = node.get(key)
        if value:
            return str(value)
    return ""


def get_date_title_variants(date_str):
    dt = datetime.strptime(date_str, "%Y-%m-%d")

    return [
        dt.strftime("%Y-%m-%d"),
        dt.strftime("%Y/%m/%d"),
        dt.strftime("%Y%m%d"),
        dt.strftime("%Y.%m.%d"),
        dt.strftime("%Y年%m月%d日"),
        f"{dt.year}年{dt.month}月{dt.day}日",
    ]

def get_project_access_user_guids(dept, dept_user_guid_map):
    """
    根据 mapping.dept 获取候选 user_guid 列表。

    返回：
    [user_guid1, user_guid2, ...]
    """

    dept = str(dept or "").strip()

    if not dept:
        return []

    user_guids = dept_user_guid_map.get(dept) or []

    if not user_guids:
        print(f"⚠️ 未找到部门对应用户: dept={dept}")
        return []

    print(
        f"✅ 匹配部门访问身份: "
        f"dept={dept}, candidate_count={len(user_guids)}"
    )

    return user_guids

def is_daily_note_title(title, target_date):
    title = title or ""
    return any(x in title for x in get_date_title_variants(target_date))


def list_folder_nodes(project_guid, folder_guid, user_guid=None):
    response = request_with_retry(
        "post",
        BASE_URL + DOC_TREE_ROUTE,
        headers=get_headers(user_guid=user_guid),
        json={
            "projectGuid": project_guid,
            "parentGuid": folder_guid,
        },
        timeout=60,
    )

    response.raise_for_status()
    response_json = response.json()

    return response_json.get("data") or []


def find_today_daily_note(project_guid, folder_guid, target_date, project_name="", user_guid=None,):
    print(
        f"🔎 查找日报: project_guid={project_guid}, "
        f"folder_guid={folder_guid}, date={target_date}, project_name={project_name}"
    )

    nodes = list_folder_nodes(
    project_guid,
    folder_guid,
    user_guid=user_guid,
    )

    for node in nodes:
        title = get_tree_node_title(node)
        guid = get_tree_node_guid(node)

        if not title or not guid:
            continue

        if not is_daily_note_title(title, target_date):
            continue

        print(f"✅ 命中日报: {title} | note_guid={guid}")

        return {
            "note_guid": guid,
            "note_title": title,
            "node": node,
        }

    return None

def try_get_daily_note_and_doc(
    project_guid,
    folder_guid,
    target_date,
    project_name,
    candidate_user_guids,
):
    """
    轮询多个 user_guid：
    谁能成功读取日报，就返回谁。
    """

    for user_guid in candidate_user_guids:
        print(
            f"🔐 尝试项目访问身份: "
            f"user_guid={user_guid}, project={project_name}"
        )

        try:
            daily_note = find_today_daily_note(
                project_guid=project_guid,
                folder_guid=folder_guid,
                target_date=target_date,
                project_name=project_name,
                user_guid=user_guid,
            )

            if not daily_note:
                print(
                    f"ℹ️ 当前用户未找到日报: "
                    f"user_guid={user_guid}"
                )
                continue

            note_guid = daily_note["note_guid"]

            # 真正读取一次文档，确认有权限
            doc_json = get_doc_json(
                note_guid,
                user_guid=user_guid,
            )

            print(
                f"✅ 项目访问身份验证成功: "
                f"user_guid={user_guid}, note={note_guid}"
            )

            return {
                "daily_note": daily_note,
                "doc_json": doc_json,
                "access_user_guid": user_guid,
            }

        except Exception as e:
            print(
                f"⚠️ 项目访问失败: "
                f"user_guid={user_guid}, error={e}"
            )

            continue

    return None

# =============================================================================
# note json
# =============================================================================


def get_doc_json(doc_id, user_guid=None):
    print(f"📄 获取日报内容: doc_id={doc_id}")

    response = request_with_retry(
        "get",
        BASE_URL + GET_DOC_API,
        headers=get_headers(user_guid=user_guid),
        params={"docId": doc_id},
        timeout=60,
    )

    response.raise_for_status()
    return response.json()


def extract_block_text(block):
    """从单个 blockContainer 提取纯文本。"""
    texts = []
    _collect_text_from_node(block, texts)
    return "".join(texts).strip()


def _collect_text_from_node(node, texts):
    """递归提取 text 节点内容，忽略 mention 等非文字节点。"""
    if isinstance(node, list):
        for item in node:
            _collect_text_from_node(item, texts)
        return

    if not isinstance(node, dict):
        return

    if node.get("type") == "text":
        text = node.get("text", "")
        if text:
            texts.append(text)

    for child in node.get("content", []) or []:
        _collect_text_from_node(child, texts)


def _collect_block_containers(node, result):
    """递归收集所有 blockContainer 节点，保持文档顺序。"""
    if isinstance(node, list):
        for item in node:
            _collect_block_containers(item, result)
        return

    if not isinstance(node, dict):
        return

    if node.get("type") == "blockContainer":
        result.append(node)
        return

    for child in node.get("content", []) or []:
        _collect_block_containers(child, result)


def _get_block_primary_type(block_container):
    """获取 blockContainer 下第一个内容节点的 type。"""
    for child in block_container.get("content", []) or []:
        child_type = child.get("type", "")
        if child_type:
            return child_type
    return ""


def _get_heading_level(block_container):
    """获取 heading 的 level 属性。"""
    for child in block_container.get("content", []) or []:
        if child.get("type") == "heading":
            attrs = child.get("attrs", {}) or {}
            level = attrs.get("level", "")
            return str(level)
    return ""


def _collect_owner_name_from_block(block_container):
    """
    从 h3 blockContainer 中提取 @人名。

    h3 可能是纯 text（如 "@汪志立"）或 mention 节点（如 label="Jerald Liu"）。
    优先从 text 中找 @前缀，若 text 无 @则尝试从 mention label 提取。
    返回去掉 @ 后的人名，或 None。
    """
    texts = []
    mention_labels = []

    def _walk(node):
        if isinstance(node, list):
            for item in node:
                _walk(item)
            return
        if not isinstance(node, dict):
            return
        if node.get("type") == "text":
            t = node.get("text", "").strip()
            if t:
                texts.append(t)
        if node.get("type") == "mention":
            attrs = node.get("attrs") or {}
            label = str(attrs.get("label") or "").strip()
            if label:
                mention_labels.append(label)
        for child in node.get("content", []) or []:
            _walk(child)

    _walk(block_container)

    full_text = " ".join(texts).strip()

    # 从 text 中找 @人名
    at_match = re.search(r"@(\S+)", full_text)
    if at_match:
        return at_match.group(1)

    # 回退到 mention label
    if mention_labels:
        return mention_labels[0]

    return None

# =============================================================================
# progress 判定
# =============================================================================


def normalize_progress_line(line):
    line = str(line or "").strip()
    line = re.sub(r"^[\-\•\*]+[\s]*", "", line).strip()
    line = re.sub(r"^\d+[.\)]\s*", "", line).strip()
    return line


def is_meaningful_progress(progress):
    if not progress:
        return False

    lines = []

    for line in progress.splitlines():
        line = normalize_progress_line(line)

        if not line:
            continue

        if line in DEFAULT_PROGRESS_VALUES:
            continue

        lines.append(line)

    return len(lines) > 0

# =============================================================================
# task parse from doc json
# =============================================================================


def parse_tasks_from_doc(doc_json):
    """
    从笔记 JSON 结构解析 Todo 任务。

    解析规则：
    1. 只有 h2 包含"平台项目重要任务进展"之后才开始抓取。
    2. h3 以 @ 开头时，提取 @后的人名作为 current_owner_name。
    3. 🚀 开头的行为新的 Todo 标题行，任务归属当前 current_owner_name。
    4. h2/h3（非@人名）为任务边界，会终结当前任务。

    兼容格式：
    - 🚀 【项目名】&【任务标题】&【50001】
    - <任务截止时间>: 2026-05-20  只展示，不回写
    - <承诺完成时间：YYYY-MM-DD>: 2026-05-21
    - <风险等级：高/中/低>: 中
    - <今日进展描述>:
      - 实际进展
    """

    content = (doc_json.get("data") or {}).get("content", [])

    all_blocks = []
    _collect_block_containers(content, all_blocks)

    tasks = []
    skip_default_count = 0

    print("🚀 开始解析日报内容（JSON 结构）")

    current_task = None
    wbs_section_started = False
    current_owner_name = None

    for block in all_blocks:
        block_text = extract_block_text(block)
        primary_type = _get_block_primary_type(block)

        # heading 处理
        if primary_type == "heading":
            level = _get_heading_level(block)

            if level == "2":
                if current_task:
                    skip_default_count = _finalize_task(
                        current_task, tasks, skip_default_count
                    )
                    current_task = None

                # h2 包含"平台项目重要任务进展"→ 开启 WBS 抓取区
                if "平台项目重要任务进展" in block_text:
                    if not wbs_section_started:
                        wbs_section_started = True
                        print("✅ 检测到 WBS 任务区域开始")
                    current_owner_name = None
                else:
                    # 其他 h2 关闭 WBS 区域
                    if wbs_section_started:
                        wbs_section_started = False
                        current_owner_name = None
                        print("ℹ️ WBS 任务区域结束")

                continue

            if level == "3":
                if current_task:
                    skip_default_count = _finalize_task(
                        current_task, tasks, skip_default_count
                    )
                    current_task = None

                # h3 以 @ 开头 → 提取 owner 人名
                owner_name = _collect_owner_name_from_block(block)
                if owner_name:
                    current_owner_name = owner_name
                    print(f"👤 检测到 owner 区域: @{current_owner_name}")
                else:
                    current_owner_name = None

                continue

        # 不在 WBS 区域内，跳过
        if not wbs_section_started:
            continue

        # 🚀 开头 = 新的 Todo 标题行
        if "🚀" in block_text:
            if current_task:
                skip_default_count = _finalize_task(
                    current_task, tasks, skip_default_count
                )

            # 取标题中的最后一个纯数字【xxx】作为 todo_id。
            # 这样即使标题/项目名中误含数字，也优先取末尾 id。
            todo_matches = re.findall(r"【(\d+)】", block_text)
            todo_id = safe_int(todo_matches[-1]) if todo_matches else None

            current_task = {
                "todo_id": todo_id,
                "title_line": block_text,
                "owner_name": current_owner_name,

                # 承诺完成时间
                "promise_raw": None,
                "promise_date": None,
                "promise_invalid": False,

                # 风险等级
                "risk_raw": None,
                "risk_level": None,
                "risk_invalid": False,

                # 今日进展
                "progress_lines": [],
                "in_progress_section": False,
            }

            continue

        # 还没遇到 🚀，跳过
        if not current_task:
            continue

        # 任务截止时间仅展示，不回写
        due_match = re.match(r"<任务截止时间[^>]*>:[ \t]*(.*)", block_text)
        if due_match:
            current_task["in_progress_section"] = False
            continue

        # 承诺完成时间：兼容 <承诺完成时间：YYYY-MM-DD>: 2026-05-21
        promise_match = re.match(r"<承诺完成时间[^>]*>:[ \t]*(.*)", block_text)
        if promise_match:
            user_value = promise_match.group(1).strip()
            current_task["promise_raw"] = user_value

            # 宽松匹配 YYYY-M-D / YYYY-MM-DD，自动补零
            date_match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", user_value)
            if date_match:
                y, m, d = date_match.groups()
                current_task["promise_date"] = f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
            elif user_value:
                current_task["promise_invalid"] = True

            current_task["in_progress_section"] = False
            continue

        # 风险等级：兼容 <风险等级：高/中/低>: 中
        risk_match = re.match(r"<风险等级[^>]*>:[ \t]*(.*)", block_text)
        if risk_match:
            user_value = risk_match.group(1).strip()
            current_task["risk_raw"] = user_value

            if user_value in RISK_MAP_CN_TO_EN:
                current_task["risk_level"] = RISK_MAP_CN_TO_EN[user_value]
            elif user_value:
                current_task["risk_invalid"] = True

            current_task["in_progress_section"] = False
            continue

        # 今日进展描述：兼容 <今日进展描述> 或 <今日进展描述：2026-05-20>
        progress_match = re.match(r"<今日进展描述[^>]*>[:：]?[ \t]*(.*)", block_text)
        if progress_match:
            current_task["in_progress_section"] = True
            inline_text = progress_match.group(1).strip()

            if inline_text:
                line = normalize_progress_line(inline_text)
                if line:
                    current_task["progress_lines"].append(line)

            continue

        # 收集进展内容（bulletListItem / paragraph）
        if current_task["in_progress_section"]:
            line = normalize_progress_line(block_text)
            if line:
                current_task["progress_lines"].append(line)

    # 保存最后一个任务
    if current_task:
        skip_default_count = _finalize_task(current_task, tasks, skip_default_count)

    print(
        f"✅ 解析完成: 有效任务={len(tasks)}, "
        f"跳过默认模板={skip_default_count}"
    )

    return tasks, skip_default_count


def _finalize_task(current_task, tasks, skip_default_count):
    """收尾一个 task，只负责从日报内容生成待匹配的解析结果。"""

    if not current_task.get("todo_id"):
        print(f"⚠️ 跳过无法解析 todo_id 的任务行: {current_task.get('title_line', '')}")
        return skip_default_count

    raw_progress = "\n".join(current_task["progress_lines"]).strip()
    has_meaningful_progress = is_meaningful_progress(raw_progress)

    if not has_meaningful_progress:
        raw_progress = ""

    warnings = []

    if current_task["promise_invalid"]:
        warnings.append("- 承诺完成时间填写格式有误，正确格式：YYYY-MM-DD")

    if current_task["risk_invalid"]:
        warnings.append("- 风险等级填写格式有误，正确格式：高/中/低")

    progress_parts = []

    if warnings:
        progress_parts.append(f"【格式检查提醒 {TARGET_DATE}】")
        progress_parts.extend(warnings)

    if has_meaningful_progress:
        if progress_parts:
            progress_parts.append("")
        progress_parts.append(raw_progress)

    progress = "\n".join(progress_parts).strip()

    if not has_meaningful_progress and not warnings:
        print(f"ℹ️ 默认模板内容，仅检查 title 前缀: todo={current_task['todo_id']}")
        skip_default_count += 1

    tasks.append({
        "id": current_task["todo_id"],
        "owner_name": current_task.get("owner_name"),

        # 用户没写或格式错误时为 None
        "promiseDate": current_task["promise_date"],
        "promiseRaw": current_task["promise_raw"],
        "promiseInvalid": current_task["promise_invalid"],

        # 用户没写或格式错误时为 None
        "riskLevel": current_task["risk_level"],
        "riskRaw": current_task["risk_raw"],
        "riskInvalid": current_task["risk_invalid"],

        "progress": progress,

        # 三字段是否有实质内容标记
        "has_progress": has_meaningful_progress,
        "has_promise": current_task["promise_date"] is not None,
        "has_risk": current_task["risk_level"] is not None,
    })

    return skip_default_count

# =============================================================================
# 与 Todo 当前状态比对，并构造更新 payload
# =============================================================================


def enrich_task_with_todo_context(parsed_task, todo_index, note_project_name=""):
    """
    用解析出的 todo_id 补充 Todo 上下文，但不在这里做任务级状态决策。

    关键原则：
    1. 只要 todo_id 匹配成功，就返回 enriched_task，即使没有进展、没有字段变化、没有 title 变化。
       因为这些“无动作”的 owner 仍然需要参与多 owner 聚合、缺失字段检查和未更新判断。
    2. progress_content 是 owner 级内容：只包含该 owner 自己的进展、格式提醒、字段一致提醒、状态变更提醒。
       李四格式错，只污染李四自己的 progress_content，不污染张三/王五。
    3. title / 未更新 / 任务级格式错误在主流程按 todo_id 聚合后统一决策。
    """

    todo_id = parsed_task["id"]
    source_todo = todo_index.get(todo_id)

    if not source_todo:
        print(
            f"⚠️ 日报中解析到 todo_id={todo_id}，"
            f"但未在 config.todo_project_id 拉取的 Todo 列表中找到，跳过更新"
        )
        return None

    owner_user_guid = get_first_owner_user_guid(source_todo)
    owner_display_name = get_first_owner_display_name(source_todo)

    matched_owner = match_owner_by_name(source_todo, parsed_task.get("owner_name"))
    if matched_owner:
        progress_user_guid = matched_owner.get("userGuid")
        progress_owner_display = (
            matched_owner.get("displayName")
            or matched_owner.get("name")
            or matched_owner.get("englishName")
            or ""
        )
    else:
        progress_user_guid = None
        progress_owner_display = ""

    source_project_name = get_todo_related_project_name(
        source_todo,
        fallback_project_name=note_project_name,
    )
    related_project = source_todo.get("relatedProject") or {}
    project_id = related_project.get("id", "")
    todo_code = get_todo_code_for_display(source_todo)

    if not owner_user_guid:
        print(
            f"⚠️ Todo 匹配成功但缺少 owners.userGuid，跳过更新: "
            f"todo_id={todo_id}, title={source_todo.get('title', '')}"
        )
        return None

    # 如果 @人名没匹配到 Todo owner，仍然保留该 parsed_task，进展用第一个 owner 身份兜底。
    # 但报告会显示为 fallback owner，日志里保留提醒。
    if not progress_user_guid:
        owner_name = parsed_task.get("owner_name") or ""
        print(
            f"⚠️ 日报 @人名「{owner_name}」未匹配到 Todo owner，"
            f"进展将使用第一个 owner 身份回写: todo_id={todo_id}"
        )
        progress_user_guid = owner_user_guid
        progress_owner_display = owner_display_name

    current_promise = normalize_date_for_compare(source_todo.get("promiseDate"))
    current_risk = str(source_todo.get("riskLevel") or "").strip()

    update_promise = parsed_task.get("promiseDate")
    update_risk = parsed_task.get("riskLevel")

    unchanged_messages = []

    if update_promise is not None:
        if normalize_date_for_compare(update_promise) == current_promise:
            unchanged_messages.append("- 承诺完成时间和上次更新一致")
            update_promise = None

    if update_risk is not None:
        if update_risk == current_risk:
            unchanged_messages.append("- 风险等级和上次更新一致")
            update_risk = None

    # owner 级格式错误。只属于当前 parsed_task 对应的 owner。
    format_errors = []
    if parsed_task.get("promiseInvalid"):
        format_errors.append({
            "field": "承诺完成时间",
            "raw": parsed_task.get("promiseRaw") or "空",
            "expected": "YYYY-MM-DD",
        })
    if parsed_task.get("riskInvalid"):
        format_errors.append({
            "field": "风险等级",
            "raw": parsed_task.get("riskRaw") or "空",
            "expected": "高/中/低",
        })

    progress_content_parts = []

    progress = parsed_task.get("progress", "").strip()
    if progress:
        progress_content_parts.append(f"【进展描述】：\n{progress}")

    change_messages = []
    if update_risk is not None:
        risk_display_new = RISK_MAP_EN_TO_CN.get(update_risk, update_risk)
        risk_display_old = RISK_MAP_EN_TO_CN.get(current_risk, current_risk) if current_risk else "无"
        change_messages.append(f"风险变更为【{risk_display_new}】（原值：{risk_display_old}）")

    if update_promise is not None:
        promise_display_old = current_promise or "无"
        change_messages.append(f"承诺完成时间变更为【{update_promise}】（原值：{promise_display_old}）")

    if change_messages:
        if progress_content_parts:
            progress_content_parts.append("")
        progress_content_parts.append(f"【状态变更提醒 {TARGET_DATE}】")
        progress_content_parts.extend([f"- {m}" for m in change_messages])

    if unchanged_messages:
        if progress_content_parts:
            progress_content_parts.append("")
        progress_content_parts.append(f"【字段一致提醒 {TARGET_DATE}】")
        progress_content_parts.extend(unchanged_messages)

    progress_content = "\n".join(progress_content_parts).strip()

    return {
        "id": todo_id,
        "type": TODO_TYPE,
        "progress_content": progress_content,
        "promiseDate": update_promise,
        "riskLevel": update_risk,
        "owner_user_guid": owner_user_guid,
        "owner_display_name": owner_display_name,
        "progress_user_guid": progress_user_guid,
        "progress_owner_display": progress_owner_display,
        "source_todo_title": source_todo.get("title", ""),
        "source_project_name": source_project_name,
        "todo_code": todo_code,
        "format_errors": format_errors,
        "currentPromiseDate": current_promise,
        "currentRiskLevel": current_risk,
        "project_id": project_id,
        "lastModifyDate": source_todo.get("lastModifyDate", ""),

        # 真实用户填写状态。title / 未更新只能看这些字段，不能看 progress_content。
        "has_progress": parsed_task.get("has_progress", False),
        "has_user_progress": parsed_task.get("has_progress", False),
        "has_promise": parsed_task.get("has_promise", False),
        "has_risk": parsed_task.get("has_risk", False),
        "parsed_owner_name": parsed_task.get("owner_name") or "",
    }

# =============================================================================
# 飞书格式错误提醒卡片
# =============================================================================


def build_format_error_card_content(task):
    """构造格式错误卡片正文。"""

    project_name = task.get("source_project_name", "未命名项目")
    todo_name = task.get("source_todo_title", "未命名任务")
    todo_code = task.get("todo_code", "")

    # 新增：项目ID
    project_id = task.get("project_id", "")

    task_line = f"【{project_name}】【{todo_name}】【{todo_code}】"

    lines = [
        f"您的任务 **{task_line}** 格式填写有误",
        "",

        # =========================
        # Markdown 表格
        # =========================
        "| 项目ID | 项目名称 | 任务Code | 任务名称 |",
        "| --- | --- | --- | --- |",
        f"| {project_id} | {project_name} | {todo_code} | {todo_name} |",

        "",
        "**错误详情：**",
    ]

    for error in task.get("format_errors", []) or []:
        field = error.get("field", "")
        raw_value = error.get("raw", "")
        expected = error.get("expected", "")

        lines.append(
            f"- **<{field}>** 填写为：**{raw_value}**，应填写格式：**{expected}**"
        )

    lines.extend([
        "",
        "请及时前往项目中心更新。",
    ])

    return "\n".join(lines)


def build_format_error_plain_text(task):
    """构造普通文本消息，作为 interactive_content 外的兜底内容。"""
    project_name = task.get("source_project_name", "未命名项目")
    todo_name = task.get("source_todo_title", "未命名任务")
    todo_code = task.get("todo_code", "")

    error_parts = []
    for error in task.get("format_errors", []) or []:
        field = error.get("field", "")
        raw_value = error.get("raw", "")
        expected = error.get("expected", "")
        error_parts.append(f"<{field}>填写为：{raw_value}，应填写格式为：{expected}")

    error_text = "；".join(error_parts)

    return (
        f"您的任务：【{project_name}】【{todo_name}】【{todo_code}】格式填写有误。"
        f"{error_text}。请前往项目中心及时更新：{FORMAT_ERROR_BUTTON_URL}"
    )


def build_format_error_feishu_card(task):
    """
    构造飞书卡片。
    参考用户提供的 schema 2.0 卡片结构，但这里只保留一个"前往项目中心"按钮。
    """
    card_content = build_format_error_card_content(task)

    return {
        "schema": "2.0",
        "header": {
            "padding": "12px 8px 12px 8px",
            "template": "red",
            "title": {
                "content": FORMAT_ERROR_CARD_TITLE,
                "tag": "plain_text",
            },
        },
        "body": {
            "vertical_spacing": "12px",
            "elements": [
                {
                    "tag": "markdown",
                    "content": card_content,
                    "margin": "0px",
                    "text_size": "normal",
                },
                {
                    "tag": "column_set",
                    "flex_mode": "stretch",
                    "horizontal_spacing": "8px",
                    "margin": "0px",
                    "columns": [
                        {
                            "tag": "column",
                            "width": "weighted",
                            "weight": 1,
                            "elements": [
                                {
                                    "tag": "button",
                                    "type": "primary_filled",
                                    "width": "fill",
                                    "margin": "4px 0px 4px 0px",
                                    "text": {
                                        "tag": "plain_text",
                                        "content": "前往项目中心",
                                    },
                                    "behaviors": [
                                        {
                                            "type": "open_url",
                                            "default_url": FORMAT_ERROR_BUTTON_URL,
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                },
            ],
        },
    }


def send_format_error_card(task):
    """
    给 @人名匹配到的 owner（progress_user_guid）发送格式错误提醒卡片。

    v4 变更：发送对象从 Todo 第一个 owner 改为 @人名匹配到的 owner。
    通过 send_message_api（BASE_URL + MESSAGE_SEND_ROUTE）发送。
    发送失败不中断 Todo 回写。
    """
    format_errors = task.get("format_errors", []) or []
    if not format_errors:
        return False

    # v4：发送给 @人名匹配到的 owner
    receiver_guid = task.get("progress_user_guid")
    if not receiver_guid:
        print(f"⚠️ 格式错误卡片未发送：缺少 receiver progress_user_guid, todo={task.get('id')}")
        return False

    card = build_format_error_feishu_card(task)
    plain_text = build_format_error_plain_text(task)

    print(
        f"📩 准备发送格式错误提醒卡片: "
        f"todo={task.get('id')}, receiver={receiver_guid}"
    )

    try:
        response = send_message_api(
            receiver_guids=[receiver_guid],
            title=FORMAT_ERROR_CARD_TITLE,
            content=plain_text,
            sender_guid=FORMAT_ERROR_SENDER_GUID,
            interactive_content=card,
        )

        if hasattr(response, "status_code"):
            ok = response.status_code == 200
            try:
                response_json = response.json()
            except Exception:
                response_json = {}

            if ok and (response_json.get("data") or response_json.get("code") in (0, 200, "0", "200", None)):
                print(f"✅ 格式错误提醒卡片发送成功: todo={task.get('id')}")
                return True

            print(f"❌ 格式错误提醒卡片发送失败: status={response.status_code}, text={getattr(response, 'text', '')}")
            return False

        print(f"✅ 格式错误提醒卡片发送完成: todo={task.get('id')}, response={response}")
        return True

    except Exception as e:
        print(f"❌ 发送格式错误提醒卡片异常: todo={task.get('id')}, error={e}")
        print(traceback.format_exc())
        return False


# =============================================================================
# 报告生成
# =============================================================================


def escape_html(text):
    text = str(text or "")
    return (
        text
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def create_workspace_doc(project_guid, parent_id, title, tag_list=None):
    """在指定项目和文件夹下创建一个新的空白文档。"""
    print(f"📄 创建文档: title={title}, project_guid={project_guid}, parent_id={parent_id}")

    payload = {
        "categoryName": title,
        "projectGuid": project_guid,
        "parentId": parent_id,
        "categoryType": 2,
        "tagList": tag_list or []
    }

    response = request_with_retry(
        "post",
        BASE_URL + WORKSPACE_SAVE_ROUTE,
        headers=get_headers(),
        json=payload,
        timeout=30
    )

    if response.status_code != 200:
        raise Exception(
            f"创建文档失败: status={response.status_code}, text={response.text[:500]}"
        )

    response_json = response.json()
    doc_guid = response_json.get("data")

    if not doc_guid:
        raise Exception(f"创建文档返回无效 data: response={response_json}")

    print(f"✅ 文档创建成功: title={title}, doc_guid={doc_guid}")
    return doc_guid


def insert_markdown_to_note(note_guid, markdown_content, writer_user_guid=None):
    if not markdown_content.strip():
        print(f"ℹ️ note={note_guid} 无可写入内容，跳过")
        return None

    print(
        f"📝 写入报告: note_guid={note_guid}, "
        f"writer_user_guid={writer_user_guid or USER_GUID}"
    )

    payload = {
        "note_guid": note_guid,
        "markdown_content": markdown_content,
        "mode": "a",
        "location": 1
    }

    response = request_with_retry(
        "post",
        BASE_URL + MD_INSERT_ROUTE,
        headers=get_headers(user_guid=writer_user_guid),
        json=payload,
        timeout=60
    )

    if response.status_code != 200:
        raise Exception(
            f"写入笔记失败: "
            f"note_guid={note_guid}, "
            f"status={response.status_code}, "
            f"text={response.text[:500]}"
        )

    response_json = response.json()

    print(f"✅ 写入成功: note_guid={note_guid}")

    return response_json



def _report_owner_key(record):
    """报告统计用人员 key。任务整体记录没有具体人员，使用 owners_display 展开。"""
    owner = str(record.get("owner_name") or "").strip()
    dept = str(record.get("owner_dept") or "").strip()
    if not owner or owner == "任务整体":
        return ""
    return f"{owner}({dept})" if dept else owner


def _split_owner_display_names(text):
    """将 owners_display 里的顿号分隔人员展开，用于卡片简表。"""
    text = str(text or "").strip()
    if not text:
        return []
    names = []
    for part in re.split(r"[、,，;；]", text):
        part = part.strip()
        if part and part not in names:
            names.append(part)
    return names


def _uniq_keep_order(items):
    result = []
    seen = set()
    for item in items:
        item = str(item or "").strip()
        if not item or item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result


def _format_people_for_card(people, limit=12):
    people = _uniq_keep_order(people)
    if not people:
        return "无"
    if len(people) <= limit:
        return "、".join(people)
    return "、".join(people[:limit]) + f" 等{len(people)}人"


def _summarize_report_data(report_data):
    """统一计算报告与卡片的摘要口径。"""
    fail_records = report_data.get("fail_records", []) or []

    task_not_updated = [r for r in fail_records if r.get("fail_type") == "not_updated"]
    person_not_updated = [r for r in fail_records if r.get("fail_type") == "missing_fields"]
    format_error = [r for r in fail_records if r.get("fail_type") == "format_error"]
    exception_records = [
        r for r in fail_records
        if r.get("fail_type") not in ("not_updated", "missing_fields", "format_error")
    ]

    captured_owners = report_data.get("captured_owners", set()) or set()
    today_update_owners = report_data.get("today_update_owners", set()) or set()

    person_not_updated_people = _uniq_keep_order(_report_owner_key(r) for r in person_not_updated)
    format_error_people = _uniq_keep_order(_report_owner_key(r) for r in format_error)
    exception_people = _uniq_keep_order(_report_owner_key(r) for r in exception_records)

    task_not_updated_people = []
    for r in task_not_updated:
        task_not_updated_people.extend(_split_owner_display_names(r.get("owners_display", "")))
    task_not_updated_people = _uniq_keep_order(task_not_updated_people)

    return {
        "task_not_updated": task_not_updated,
        "person_not_updated": person_not_updated,
        "format_error": format_error,
        "exception_records": exception_records,
        "captured_owners": _uniq_keep_order(captured_owners),
        "today_update_owners": _uniq_keep_order(today_update_owners),
        "person_not_updated_people": person_not_updated_people,
        "format_error_people": format_error_people,
        "exception_people": exception_people,
        "task_not_updated_people": task_not_updated_people,
    }


def build_report_markdown(target_date, report_data):
    """构建报告的 Markdown/HTML 内容。报告正文保留详细表格。"""
    total_tasks = report_data.get("total_parsed_tasks", 0)
    captured_records = report_data.get("total_update_tasks", 0)
    today_update_records = report_data.get("today_update_count", 0)
    summary = _summarize_report_data(report_data)

    task_not_updated = summary["task_not_updated"]
    person_not_updated = summary["person_not_updated"]
    format_error = summary["format_error"]
    exception_records = summary["exception_records"]

    lines = []
    lines.append("<h2>任务更新情况汇总</h2>")
    lines.append(f"<p><strong>总解析任务</strong>：{total_tasks}</p>")
    lines.append(f"<p><strong>成功捕捉更新</strong>：{captured_records} 条记录，{len(summary['captured_owners'])} 人</p>")
    lines.append(f"<p><strong>今日更新</strong>：{today_update_records} 条记录，{len(summary['today_update_owners'])} 人</p>")
    lines.append(f"<p><strong>未更新</strong>：{len(person_not_updated)} 条记录，{len(summary['person_not_updated_people'])} 人</p>")
    lines.append(f"<p><strong>格式错误</strong>：{len(format_error)} 条记录，{len(summary['format_error_people'])} 人</p>")
    lines.append(f"<p><strong>异常/缺失记录</strong>：{len(exception_records)} 条，涉及 {len(summary['exception_people'])} 人</p>")
    lines.append("")

    if task_not_updated:
        lines.append("<h2>任务未更新状态汇总</h2>")
        lines.append("<p>说明：只要任务下存在任一 owner 三项未完整填写，任务 title 即保持【X次未更新】；X 为任务级展示次数，不代表个人级更新时间。</p>")
        lines.append("<table><thead><tr><th>任务Code</th><th>任务名称</th><th>Title未更新次数</th><th>未完整更新负责人</th></tr></thead><tbody>")
        for r in task_not_updated:
            code = escape_html(r.get("todo_code", ""))
            title = escape_html(r.get("task_title", ""))
            days = escape_html(str(r.get("days_not_updated", "")))
            owners_display = escape_html(r.get("owners_display", "") or r.get("owner_name", ""))
            lines.append(f"<tr><td>{code}</td><td>{title}</td><td>{days}次未更新</td><td>{owners_display}</td></tr>")
        lines.append("</tbody></table>")
        lines.append("")

    if person_not_updated:
        lines.append("<h2>个人未更新情况</h2>")
        lines.append("<p>说明：个人未更新指进展描述、承诺完成时间、风险等级三项中任一缺失或格式非法。已完整填写的人不会因为同任务其他 owner 未更新而进入本表。</p>")
        lines.append("<table><thead><tr><th>负责人(部门)</th><th>任务Code</th><th>任务名称</th><th>Title未更新次数</th><th>未更新原因</th></tr></thead><tbody>")
        for r in person_not_updated:
            owner = escape_html(r.get("owner_name", ""))
            dept = escape_html(r.get("owner_dept", ""))
            owner_display = f"{owner}({dept})" if dept else owner
            code = escape_html(r.get("todo_code", ""))
            title = escape_html(r.get("task_title", ""))
            days = escape_html(str(r.get("days_not_updated", "")))
            reason = escape_html(r.get("missing_fields", "") or r.get("reason", ""))
            lines.append(f"<tr><td>{owner_display}</td><td>{code}</td><td>{title}</td><td>{days}次</td><td>{reason}</td></tr>")
        lines.append("</tbody></table>")
        lines.append("")

    if format_error:
        lines.append("<h2>格式错误人员情况</h2>")
        lines.append("<table><thead><tr><th>负责人(部门)</th><th>任务Code</th><th>任务名称</th><th>错误字段</th><th>填写值</th><th>正确格式</th></tr></thead><tbody>")
        for r in format_error:
            owner = escape_html(r.get("owner_name", ""))
            dept = escape_html(r.get("owner_dept", ""))
            owner_display = f"{owner}({dept})" if dept else owner
            code = escape_html(r.get("todo_code", ""))
            title = escape_html(r.get("task_title", ""))
            field = escape_html(r.get("error_field", ""))
            raw = escape_html(r.get("error_raw", ""))
            expected = escape_html(r.get("error_expected", ""))
            lines.append(f"<tr><td>{owner_display}</td><td>{code}</td><td>{title}</td><td>{field}</td><td>{raw}</td><td>{expected}</td></tr>")
        lines.append("</tbody></table>")
        lines.append("")

    if exception_records:
        lines.append("<h2>更新异常 / 缺失记录</h2>")
        lines.append("<p>说明：这里记录接口失败、Todo 未匹配、权限异常等系统处理异常，不包含正常业务口径下的个人未更新。</p>")
        lines.append("<table><thead><tr><th>负责人(部门)</th><th>任务Code</th><th>任务名称</th><th>异常类型</th><th>原因</th></tr></thead><tbody>")
        for r in exception_records:
            owner = escape_html(r.get("owner_name", ""))
            dept = escape_html(r.get("owner_dept", ""))
            owner_display = f"{owner}({dept})" if dept else owner
            code = escape_html(r.get("todo_code", ""))
            title = escape_html(r.get("task_title", ""))
            fail_type = escape_html(r.get("fail_type", ""))
            reason = escape_html(r.get("reason", "") or r.get("missing_fields", "") or "-")
            lines.append(f"<tr><td>{owner_display}</td><td>{code}</td><td>{title}</td><td>{fail_type}</td><td>{reason}</td></tr>")
        lines.append("</tbody></table>")
        lines.append("")

    return "\n".join(lines)


def build_report_card_content(report_data):
    """构建报告卡片正文。卡片只展示概览和问题人员名单，不展示详细表格。"""
    total_tasks = report_data.get("total_parsed_tasks", 0)
    captured_records = report_data.get("total_update_tasks", 0)
    today_update_records = report_data.get("today_update_count", 0)
    summary = _summarize_report_data(report_data)

    lines = []
    lines.append(f"**总解析任务**：{total_tasks}")
    lines.append(f"**成功捕捉更新**：{captured_records} 条记录，{len(summary['captured_owners'])} 人")
    lines.append(f"**今日更新**：{today_update_records} 条记录，{len(summary['today_update_owners'])} 人")
    lines.append(f"**未更新**：{len(summary['person_not_updated'])} 条记录，{len(summary['person_not_updated_people'])} 人")
    lines.append(f"**格式错误**：{len(summary['format_error'])} 条记录，{len(summary['format_error_people'])} 人")
    lines.append(f"**异常/缺失记录**：{len(summary['exception_records'])} 条，涉及 {len(summary['exception_people'])} 人")

    issue_lines = []
    if summary["task_not_updated_people"]:
        issue_lines.append(f"**任务未更新涉及人员**：{_format_people_for_card(summary['task_not_updated_people'])}")
    if summary["person_not_updated_people"]:
        issue_lines.append(f"**个人未更新人员**：{_format_people_for_card(summary['person_not_updated_people'])}")
    if summary["format_error_people"]:
        issue_lines.append(f"**格式错误人员**：{_format_people_for_card(summary['format_error_people'])}")
    if summary["exception_people"]:
        issue_lines.append(f"**异常/缺失涉及人员**：{_format_people_for_card(summary['exception_people'])}")

    if issue_lines:
        lines.append("")
        lines.extend(issue_lines)

    return "\n".join(lines)

def build_report_feishu_card(title, card_content, note_url):
    """构造飞书卡片，包含"查看报告"按钮。"""
    return {
        "schema": "2.0",
        "header": {
            "padding": "12px 8px 12px 8px",
            "template": "blue",
            "title": {
                "content": title,
                "tag": "plain_text"
            }
        },
        "body": {
            "vertical_spacing": "12px",
            "elements": [
                {
                    "tag": "markdown",
                    "content": card_content,
                    "margin": "0px",
                    "text_size": "normal"
                },
                {
                    "tag": "column_set",
                    "flex_mode": "stretch",
                    "horizontal_spacing": "8px",
                    "margin": "0px",
                    "columns": [
                        {
                            "tag": "column",
                            "width": "weighted",
                            "weight": 1,
                            "elements": [
                                {
                                    "tag": "button",
                                    "type": "primary_filled",
                                    "width": "fill",
                                    "margin": "4px 0px 4px 0px",
                                    "text": {
                                        "tag": "plain_text",
                                        "content": "查看报告"
                                    },
                                    "behaviors": [
                                        {
                                            "type": "open_url",
                                            "default_url": note_url
                                        }
                                    ]
                                }
                            ]
                        }
                    ]
                }
            ]
        }
    }


def send_webhook_card(webhook_url, card, max_retries=3, retry_interval=5):
    """通过 Webhook 发送飞书卡片到群。"""
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.post(
                url=webhook_url,
                headers={"Content-Type": "application/json"},
                json={"msg_type": "interactive", "card": card},
                timeout=10
            )
            result = response.json()

            if result.get("code") == 0 or result.get("StatusCode") == 0:
                return result

            if attempt < max_retries:
                print(f"  -> ⚠️ Webhook 发送失败 (尝试 {attempt}/{max_retries}): {result}，{retry_interval}秒后重试...")
                time.sleep(retry_interval)
            else:
                return result
        except Exception as e:
            if attempt < max_retries:
                print(f"  -> ⚠️ Webhook 发送异常 (尝试 {attempt}/{max_retries}): {e}，{retry_interval}秒后重试...")
                time.sleep(retry_interval)
            else:
                raise

    return {"code": -1, "msg": "max retries exceeded"}


def send_report_message(report_title, doc_guid, report_data):
    """发送报告卡片消息到配置的群和个人。"""
    if not REPORT_WEBHOOK and not REPORT_SENDER_GUID:
        print("ℹ️ 未配置 report_webhook 或 report_sender_guid，跳过消息发送")
        return

    note_url = f"{BASE_URL}/workspace/{doc_guid}"
    card_content = build_report_card_content(report_data)
    card = build_report_feishu_card(report_title, card_content, note_url)

    has_sent_any = False

    if REPORT_WEBHOOK:
        for idx, url in enumerate(REPORT_WEBHOOK, 1):
            try:
                print(f"📢 正在发送群消息 (Webhook {idx}/{len(REPORT_WEBHOOK)})...")
                result = send_webhook_card(url, card)

                if result.get("code") == 0 or result.get("StatusCode") == 0:
                    print(f"  -> ✅ 群消息发送成功: {url[:30]}...")
                    has_sent_any = True
                else:
                    print(f"  -> ❌ 群消息发送失败: {result}")
            except Exception as e:
                print(f"  -> ❌ 群消息发送异常: {e}")

    if REPORT_SENDER_GUID:
        try:
            print(f"📩 正在发送个人消息给 {len(REPORT_SENDER_GUID)} 人...")
            text_content = f"【{report_title}】已生成，请点击查看。\n<a href='{note_url}'>点击查看详情</a>"
            response = send_message_api(
                receiver_guids=REPORT_SENDER_GUID,
                title=report_title,
                content=text_content,
                interactive_content=card,
            )
            if hasattr(response, "status_code") and response.status_code == 200:
                try:
                    resp_json = response.json()
                except Exception:
                    resp_json = {}
                if resp_json.get("data") or resp_json.get("code") in (0, 200, "0", "200", None):
                    print("  -> ✅ 个人消息发送成功")
                    has_sent_any = True
                else:
                    print(f"  -> ❌ 个人消息发送失败: {response.text if hasattr(response, 'text') else '无响应'}")
            else:
                print(f"  -> ❌ 个人消息发送失败: {response.text if response and hasattr(response, 'text') else '无响应'}")
        except Exception as e:
            print(f"  -> ❌ 个人消息发送异常: {e}")

    if has_sent_any:
        print(f"✅ 报告消息发送完成")
    else:
        print("⚠️ 报告消息未能成功发送到任何目标")


def write_report(target_date, report_data):
    """生成任务更新报告并写入 Workspace。"""
    if not REPORT_TARGET_PROJECT_GUID or not REPORT_UPDATE_PARENT_GUID:
        print("ℹ️ 未配置 report_target_project_guid 或 report_update_parent_guid，跳过报告生成")
        return None

    report_title = f"{target_date} 项目管理任务进展更新情况"

    print(f"\n🚀 开始生成任务更新报告: {report_title}")

    doc_guid = create_workspace_doc(
        project_guid=REPORT_TARGET_PROJECT_GUID,
        parent_id=REPORT_UPDATE_PARENT_GUID,
        title=report_title
    )

    markdown_content = build_report_markdown(target_date, report_data)

    print("\n================ 报告内容预览 ================\n")
    print(markdown_content)
    print("\n==============================================\n")

    insert_markdown_to_note(
        note_guid=doc_guid,
        markdown_content=markdown_content
    )

    print(f"✅ 任务更新报告写入完成: {report_title}")

    return doc_guid, report_title


# =============================================================================
# title 前缀处理
# =============================================================================

_TITLE_PREFIX_RE = re.compile(r"^【\d+次未更新】|^【格式错误】")


def strip_title_prefixes(title):
    """去掉 title 中的【X次未更新】【格式错误】前缀，恢复原始标题。"""
    title = title or ""
    while True:
        new_title = _TITLE_PREFIX_RE.sub("", title)
        if new_title == title:
            break
        title = new_title
    return title


def calc_days_since_modify(last_modify_date_str, now=None):
    """
    根据 lastModifyDate 计算距今的工作日天数（周六周日不计入）。
    lastModifyDate 格式：2026-05-24 15:08:54

    规则：
    - 只数周一~周五，周六周日跳过。
    - 当天不满 24h 不算一天（即 modify 当天本身不计入）。
    - 例如：周五 10:00 修改 → 下周一 10:00 = 1天（周六周日不算）。
    """
    if not last_modify_date_str:
        return 0

    try:
        last_dt = datetime.strptime(
            str(last_modify_date_str).strip()[:19], "%Y-%m-%d %H:%M:%S"
        )
    except Exception:
        return 0

    ref = now or datetime.now()

    workdays = 0
    check_date = last_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = ref.replace(hour=0, minute=0, second=0, microsecond=0)
    one_day = timedelta(days=1)

    # 从 modify 的下一天开始算（modify 当天本身不计入）
    check_date += one_day

    while check_date <= end_date:
        if check_date.weekday() < 5:  # 0=Mon ... 5=Sat, 6=Sun
            workdays += 1
        check_date += one_day

    return workdays


def build_update_title(task):
    """
    根据已经决策好的任务级状态生成 Todo title。

    v6 口径：
    1. 个人未更新 = 进展描述 / 承诺完成时间 / 风险等级 三项中任一缺失或格式非法。
    2. 任务未更新 = 该任务下存在任意 owner 处于个人未更新状态。
    3. title 是任务级状态：只要存在未完整更新 owner，就添加【X次未更新】。
    4. X 来自任务级 lastModifyDate 计算结果；若为 0，为避免出现【0次未更新】，展示为 1。
       其含义是：当前回收周期内该任务仍存在未完整更新 owner。
    5. 只有所有 owner 都完整更新，才清除【X次未更新】。
    6. 格式错误独立叠加【格式错误】。
    """
    base_title = strip_title_prefixes(task.get("source_todo_title", ""))
    current_title = task.get("source_todo_title", "")

    has_incomplete_owner = bool(task.get("has_incomplete_owner", False))
    has_format_error = bool(task.get("format_errors"))

    if "days_not_updated" in task:
        days_not_updated = int(task.get("days_not_updated") or 0)
    else:
        days_not_updated = calc_days_since_modify(task.get("lastModifyDate", ""))

    # 展示用次数：业务要求 title 仍使用【X次未更新】，但不展示 0。
    display_not_updated_count = int(task.get("display_not_updated_count") or 0)
    if has_incomplete_owner and display_not_updated_count <= 0:
        display_not_updated_count = max(1, days_not_updated)

    prefixes = []

    if has_incomplete_owner:
        prefixes.append(f"【{display_not_updated_count}次未更新】")

    if has_format_error:
        prefixes.append("【格式错误】")

    if not prefixes:
        if current_title != base_title:
            return base_title
        return None

    new_title = "".join(prefixes) + base_title

    if new_title == current_title:
        return None

    return new_title


# =============================================================================
# update todo
# =============================================================================


def update_todo(task):
    """更新 promiseDate / riskLevel / title，使用 task 中指定的 owner_user_guid 身份。进展内容走 insertProgress 接口。"""
    payload = {
        "type": TODO_TYPE,
        "id": task["id"],
    }

    # 没写 / 与当前值一致 / 格式错误 -> None，不传入接口
    if task.get("promiseDate") is not None:
        payload["promiseDate"] = task["promiseDate"]

    if task.get("riskLevel") is not None:
        payload["riskLevel"] = task["riskLevel"]

    # title 前缀：由上游统一决策，update_todo 只执行，不再重复计算。
    new_title = task.get("new_title")
    if new_title is not None:
        payload["title"] = new_title

    # 无字段需要更新则跳过
    if "promiseDate" not in payload and "riskLevel" not in payload and "title" not in payload:
        print(f"ℹ️ 无字段需更新，跳过 update: id={task['id']}")
        return None

    # v4：字段更新使用第一个 owner 身份
    owner_user_guid = task.get("owner_user_guid")
    owner_display_name = task.get("owner_display_name", "")

    print(
        f"\n🚀 更新 Todo 字段: id={task['id']}, "
        f"owner={owner_display_name}, owner_user_guid={owner_user_guid}"
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))

    response = request_with_retry(
        "post",
        TODO_BASE_URL + TODO_UPDATE_API,
        headers=get_todo_headers(user_guid=owner_user_guid),
        json=payload,
        timeout=30,
    )

    response.raise_for_status()
    result = response.json()

    if result.get("code") not in (0, 200, "0", "200", None):
        raise Exception(f"Todo 更新接口返回异常: response={result}")

    print(f"✅ 更新成功: todo={task['id']}")
    return result


def insert_todo_progress(task):
    """新增进展，使用 @人名匹配到的 owner（progress_user_guid）身份回写。"""
    payload = {
        "id": task["id"],
        "type": TODO_TYPE,
        "content": task["progress_content"],
    }

    # v4：进展回写使用 @人名匹配到的 owner 身份
    progress_user_guid = task.get("progress_user_guid")

    print(
        f"\n🚀 新增进展: id={task['id']}, "
        f"progress_user_guid={progress_user_guid}"
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))

    response = request_with_retry(
        "post",
        TODO_BASE_URL + TODO_INSERT_PROGRESS_API,
        headers=get_todo_headers(user_guid=progress_user_guid),
        json=payload,
        timeout=30,
    )

    response.raise_for_status()
    result = response.json()

    if result.get("code") not in (0, 200, "0", "200", None):
        raise Exception(f"新增进展接口返回异常: response={result}")

    print(f"✅ 新增进展成功: todo={task['id']}")
    return result



def _owner_display_name(owner):
    return str(
        owner.get("displayName")
        or owner.get("name")
        or owner.get("englishName")
        or owner.get("loginName")
        or ""
    ).strip()


def _owner_name_candidates(owner):
    return {
        str(owner.get("name") or "").strip(),
        str(owner.get("englishName") or "").strip(),
        str(owner.get("displayName") or "").strip(),
    } - {""}


def _build_owner_incomplete_reasons(task):
    """
    owner 级未更新原因。
    三项任一缺失或非法，即视为该 owner 未完整更新。
    """
    reasons = []
    format_error_fields = {
        str(err.get("field") or "").strip()
        for err in (task.get("format_errors") or [])
    }

    if not task.get("has_progress"):
        reasons.append("缺少进展描述")

    if "承诺完成时间" in format_error_fields:
        reasons.append("承诺完成时间格式错误")
    elif not task.get("has_promise"):
        reasons.append("缺少承诺完成时间")

    if "风险等级" in format_error_fields:
        reasons.append("风险等级格式错误")
    elif not task.get("has_risk"):
        reasons.append("缺少风险等级")

    return reasons


def _get_task_not_updated_count(last_modify_date):
    """
    title/report 展示用的任务级未更新次数。
    lastModifyDate 是任务级字段；当存在未完整更新 owner 但计算值为 0 时，展示为 1，避免【0次未更新】。
    """
    raw_days = calc_days_since_modify(last_modify_date)
    return raw_days, max(1, raw_days)

# =============================================================================
# 主流程
# =============================================================================

print("\n==========================================")
print("🚀 Todo 日报回写应用启动 (v6 closed-loop)")
print("==========================================\n")

print(f"BASE_URL: {BASE_URL}")
print(f"TODO_BASE_URL: {TODO_BASE_URL}")
print(f"TARGET_DATE: {TARGET_DATE}")
print(f"USER_GUID: {USER_GUID}")
print(f"MAPPING_GUID: {MAPPING_GUID}")
print(f"TODO_PROJECT_ID: {TODO_PROJECT_ID}")

mapping_rows = load_mapping_from_guid(MAPPING_GUID)

todo_list = fetch_todo_list()
if not todo_list:
    print("ℹ️ 当前未获取到 Todo 数据，流程结束")
    sys.exit(0)

todo_index = build_todo_index(todo_list)
dept_user_guid_map = build_dept_user_guid_map(todo_list)

visited_notes = set()

total_projects = 0
hit_notes = 0
total_parsed_tasks = 0
total_update_tasks = 0
progress_success_count = 0
progress_fail_count = 0
update_success_count = 0
update_fail_count = 0
skip_default_total = 0
skip_no_match_count = 0
skip_no_change_count = 0
card_success_count = 0
card_fail_count = 0
card_skip_count = 0
success_owners = set()
captured_owners = set()
today_update_count = 0
today_update_owners = set()
fail_records = []

def _get_owner_dept(task):
    """从 task 关联的 source_todo 中获取 progress_user_guid 对应的部门。"""
    src_todo = todo_index.get(task.get("id"))
    if src_todo:
        for o in src_todo.get("owners", []) or []:
            if o.get("userGuid") == task.get("progress_user_guid"):
                return o.get("fullDeptName", "")
    return ""


def _build_format_error_from_parsed(parsed_task, src_todo, owner_name, owner_dept):
    """从 parsed_task 中提取格式错误记录（enrich 返回 None 时使用）。"""
    records = []
    todo_code = get_todo_code_for_display(src_todo)
    task_title = src_todo.get("title", "")

    if parsed_task.get("promiseInvalid"):
        records.append({
            "fail_type": "format_error",
            "owner_name": owner_name,
            "owner_dept": owner_dept,
            "todo_code": todo_code,
            "task_title": task_title,
            "error_field": "承诺完成时间",
            "error_raw": parsed_task.get("promiseRaw") or "空",
            "error_expected": "YYYY-MM-DD",
        })

    if parsed_task.get("riskInvalid"):
        records.append({
            "fail_type": "format_error",
            "owner_name": owner_name,
            "owner_dept": owner_dept,
            "todo_code": todo_code,
            "task_title": task_title,
            "error_field": "风险等级",
            "error_raw": parsed_task.get("riskRaw") or "空",
            "error_expected": "高/中/低",
        })

    return records

for row in mapping_rows:
    dept = row.get("dept", "").strip()
    project_guid = row.get("project_guid")
    folder_guid = row.get("work_log_folder_guid")
    project_name = row.get("project_name", "")

    if not project_guid:
        continue

    if not folder_guid:
        continue

    total_projects += 1

    print(f"\n🚀 处理项目: {project_name}")

    candidate_user_guids = get_project_access_user_guids(
        dept,
        dept_user_guid_map,
    )

    if not candidate_user_guids:
        print(f"⚠️ 项目缺少可访问身份，跳过: {project_name}")
        continue

    access_result = try_get_daily_note_and_doc(
        project_guid=project_guid,
        folder_guid=folder_guid,
        target_date=TARGET_DATE,
        project_name=project_name,
        candidate_user_guids=candidate_user_guids,
    )

    if not access_result:
        print(f"⚠️ 所有候选用户均无法访问项目日报: {project_name}")
        continue

    daily_note = access_result["daily_note"]
    doc_json = access_result["doc_json"]

    note_guid = daily_note["note_guid"]

    if note_guid in visited_notes:
        print(f"ℹ️ 日报已处理过，跳过重复 note: {note_guid}")
        continue

    visited_notes.add(note_guid)
    hit_notes += 1

    print(
        f"✅ 命中日报: "
        f"{daily_note['note_title']}"
    )

    parsed_tasks, skip_default_count = parse_tasks_from_doc(doc_json)
    skip_default_total += skip_default_count

    if not parsed_tasks:
        print("ℹ️ 当前日报无有效 Todo 更新")
        continue

    total_parsed_tasks += len(parsed_tasks)

    print("\n======================")
    print("日报解析结果")
    print("======================")
    print(json.dumps(parsed_tasks, ensure_ascii=False, indent=2))

    # 按 todo_id 分组，同一 todo_id 可能出现在多个 owner 区域下
    tasks_by_todo_id = OrderedDict()
    for parsed_task in parsed_tasks:
        tid = parsed_task.get("id")
        if tid not in tasks_by_todo_id:
            tasks_by_todo_id[tid] = []
        tasks_by_todo_id[tid].append(parsed_task)

    # 对每个 todo_id 组分别处理：在这里统一生成任务级决策。
    # 重要：enrich_task_with_todo_context 不再因为“无动作”提前过滤 owner，
    # 否则多 owner 场景下会丢失未填写/缺字段信息。
    for tid, group_tasks in tasks_by_todo_id.items():
        source_todo = todo_index.get(tid)

        if not source_todo:
            skip_no_match_count += len(group_tasks)
            print(f"⚠️ todo_id={tid} 未在 Todo 索引中找到，跳过该组")
            continue

        owners = source_todo.get("owners", []) or []
        todo_code = get_todo_code_for_display(source_todo)
        source_title = source_todo.get("title", "")

        enriched_tasks = []
        for parsed_task in group_tasks:
            enriched_task = enrich_task_with_todo_context(
                parsed_task,
                todo_index,
                note_project_name=project_name,
            )
            if not enriched_task:
                skip_no_change_count += 1
                continue
            enriched_tasks.append(enriched_task)

        if not enriched_tasks:
            continue

        total_update_tasks += len(enriched_tasks)

        print(f"\n======================")
        print(f"待处理 Todo (id={tid}): {len(enriched_tasks)} 条 owner 记录")
        print("======================")
        print(json.dumps(enriched_tasks, ensure_ascii=False, indent=2))

        # ------------------------------------------------------------------
        # A. owner 级动作：格式错误卡片 / insertProgress / 个人未更新记录
        # ------------------------------------------------------------------
        # 任务级 lastModifyDate 只算一次；展示用次数不允许为 0。
        raw_days_not_updated, display_not_updated_count = _get_task_not_updated_count(
            source_todo.get("lastModifyDate", "")
        )

        updated_owner_guids = set()
        incomplete_owner_labels = []

        for task in enriched_tasks:
            if task.get("progress_user_guid"):
                updated_owner_guids.add(task.get("progress_user_guid"))

            owner_name = task.get("progress_owner_display") or task.get("owner_display_name") or ""
            owner_dept = _get_owner_dept(task)
            owner_label = f"{owner_name}({owner_dept})" if owner_dept else owner_name
            if owner_label:
                captured_owners.add(owner_label)

            # A1. 格式错误：owner 级，只发给/记录填错的人。
            if task.get("format_errors"):
                sent = send_format_error_card(task)
                if sent:
                    card_success_count += 1
                else:
                    card_fail_count += 1

                for err in task.get("format_errors", []):
                    fail_records.append({
                        "fail_type": "format_error",
                        "owner_name": owner_name,
                        "owner_dept": owner_dept,
                        "todo_code": task.get("todo_code", ""),
                        "task_title": task.get("source_todo_title", ""),
                        "error_field": err.get("field", ""),
                        "error_raw": err.get("raw", ""),
                        "error_expected": err.get("expected", ""),
                        "days_not_updated": display_not_updated_count,
                    })
            else:
                card_skip_count += 1

            # A2. 进展内容：owner 级，只写当前 owner 的内容。
            # progress_content 可能包含：真实进展、当前 owner 的格式提醒、字段一致提醒、状态变更提醒。
            # 但它不参与 title/未更新判断。
            if task.get("progress_content"):
                try:
                    insert_todo_progress(task)
                    progress_success_count += 1
                    success_owners.add(owner_label)
                except Exception as e:
                    progress_fail_count += 1
                    print(f"❌ 新增进展失败 todo={task['id']} error={e}")
                    print(traceback.format_exc())
                    fail_records.append({
                        "fail_type": "progress_fail",
                        "owner_name": owner_name,
                        "owner_dept": owner_dept,
                        "todo_code": task.get("todo_code", ""),
                        "task_title": task.get("source_todo_title", ""),
                        "reason": f"新增进展失败: {e}",
                    })

            # A3. 个人未更新/更新不完整：owner 级。
            # 业务口径：进展描述 / 承诺完成时间 / 风险等级 三项任一缺失或非法，即为个人未更新。
            incomplete_reasons = _build_owner_incomplete_reasons(task)
            if incomplete_reasons:
                incomplete_owner_labels.append(owner_label or owner_name or "-")
                fail_records.append({
                    "fail_type": "missing_fields",
                    "owner_name": owner_name,
                    "owner_dept": owner_dept,
                    "todo_code": task.get("todo_code", ""),
                    "task_title": task.get("source_todo_title", ""),
                    "missing_fields": "、".join(incomplete_reasons),
                    "days_not_updated": display_not_updated_count,
                })
            else:
                # 今日更新：该 owner 在该任务下三项均完整且格式合法。
                today_update_count += 1
                if owner_label:
                    today_update_owners.add(owner_label)

        # A4. Todo owners 中完全没有出现在日报解析结果里的 owner，记录为“个人未更新”。
        for owner in owners:
            owner_guid = str(owner.get("userGuid") or "").strip()
            if not owner_guid or owner_guid in updated_owner_guids:
                continue

            owner_name = _owner_display_name(owner)
            owner_dept = str(owner.get("fullDeptName") or "").strip()
            owner_label = f"{owner_name}({owner_dept})" if owner_dept else owner_name
            incomplete_owner_labels.append(owner_label or owner_name or "-")
            fail_records.append({
                "fail_type": "missing_fields",
                "owner_name": owner_name,
                "owner_dept": owner_dept,
                "todo_code": todo_code,
                "task_title": source_title,
                "missing_fields": "未在日报中填写该任务",
                "days_not_updated": display_not_updated_count,
            })

        # ------------------------------------------------------------------
        # B. 任务级状态：title / 任务未更新状态 / 任务级格式错误
        # ------------------------------------------------------------------
        all_format_errors = []
        for et in enriched_tasks:
            all_format_errors.extend(et.get("format_errors") or [])

        # v6：任务未更新 = 存在任意 owner 未完整更新。
        has_incomplete_owner = len(incomplete_owner_labels) > 0

        title_task_info = {
            "source_todo_title": source_title,
            "lastModifyDate": source_todo.get("lastModifyDate", ""),
            "format_errors": all_format_errors,
            "has_incomplete_owner": has_incomplete_owner,
            "days_not_updated": raw_days_not_updated,
            "display_not_updated_count": display_not_updated_count,
        }
        new_title = build_update_title(title_task_info)

        if has_incomplete_owner:
            # 任务级汇总只列未完整更新的人；已完整更新的 owner 不进入该报告。
            owners_display = "、".join([x for x in incomplete_owner_labels if x]) or "-"
            fail_records.append({
                "fail_type": "not_updated",
                "owner_name": "任务整体",
                "owner_dept": "",
                "owners_display": owners_display,
                "todo_code": todo_code,
                "task_title": source_title,
                "days_not_updated": display_not_updated_count,
            })

        # ------------------------------------------------------------------
        # C. 字段更新：promiseDate/riskLevel 按 Todo owners 顺序顺延。
        # ------------------------------------------------------------------
        final_promise = None
        promise_contributor = None
        final_risk = None
        risk_contributor = None

        for owner in owners:
            owner_name_val = str(owner.get("name") or "").strip()
            owner_english_name = str(owner.get("englishName") or "").strip()
            owner_display_name = str(owner.get("displayName") or "").strip()

            matched_pt = None
            for pt in group_tasks:
                pt_name = pt.get("owner_name") or ""
                if pt_name and pt_name in {owner_name_val, owner_english_name, owner_display_name}:
                    matched_pt = pt
                    break

            if matched_pt:
                if final_promise is None and matched_pt.get("promiseDate") is not None:
                    final_promise = matched_pt.get("promiseDate")
                    promise_contributor = owner
                if final_risk is None and matched_pt.get("riskLevel") is not None:
                    final_risk = matched_pt.get("riskLevel")
                    risk_contributor = owner

            if final_promise is not None and final_risk is not None:
                break

        current_promise = normalize_date_for_compare(source_todo.get("promiseDate"))
        current_risk_val = str(source_todo.get("riskLevel") or "").strip()

        if final_promise is not None and normalize_date_for_compare(final_promise) == current_promise:
            final_promise = None
            promise_contributor = None
        if final_risk is not None and final_risk == current_risk_val:
            final_risk = None
            risk_contributor = None

        owner_fields_map = OrderedDict()

        if final_promise is not None and promise_contributor:
            pg = str(promise_contributor.get("userGuid") or "").strip()
            if pg:
                owner_fields_map.setdefault(pg, {
                    "owner_obj": promise_contributor,
                    "promiseDate": None,
                    "riskLevel": None,
                })
                owner_fields_map[pg]["promiseDate"] = final_promise

        if final_risk is not None and risk_contributor:
            rg = str(risk_contributor.get("userGuid") or "").strip()
            if rg:
                owner_fields_map.setdefault(rg, {
                    "owner_obj": risk_contributor,
                    "promiseDate": None,
                    "riskLevel": None,
                })
                owner_fields_map[rg]["riskLevel"] = final_risk

        # ------------------------------------------------------------------
        # D. 执行 update：字段和 title 都消费上面已经决策好的结果，不再重算。
        # ------------------------------------------------------------------
        has_update_action = bool(owner_fields_map) or new_title is not None
        if not has_update_action:
            continue

        first_update = True
        for user_guid, fields in owner_fields_map.items():
            owner_obj = fields["owner_obj"]
            owner_display = str(owner_obj.get("displayName") or owner_obj.get("name") or owner_obj.get("englishName") or "").strip()

            update_task = {
                "id": tid,
                "type": TODO_TYPE,
                "promiseDate": fields.get("promiseDate"),
                "riskLevel": fields.get("riskLevel"),
                "owner_user_guid": user_guid,
                "owner_display_name": owner_display,
                "source_todo_title": source_title,
                "lastModifyDate": source_todo.get("lastModifyDate", ""),
                "days_not_updated": raw_days_not_updated,
                "display_not_updated_count": display_not_updated_count,
                "has_incomplete_owner": has_incomplete_owner,
                "format_errors": all_format_errors if first_update else [],
                "new_title": new_title if first_update else None,
            }

            try:
                result = update_todo(update_task)
                if result is not None:
                    update_success_count += 1
            except Exception as e:
                update_fail_count += 1
                print(f"❌ 更新失败 todo={tid} error={e}")
                print(traceback.format_exc())
                fail_records.append({
                    "fail_type": "update_fail",
                    "owner_name": owner_display,
                    "owner_dept": str((owner_obj or {}).get("fullDeptName", "")),
                    "todo_code": todo_code,
                    "task_title": source_title,
                    "reason": f"字段更新失败: {e}",
                })

            first_update = False

        # 如果没有字段更新，但 title 需要更新，使用第一位 owner 身份单独更新 title。
        if not owner_fields_map and new_title is not None:
            fallback_guid = get_first_owner_user_guid(source_todo)
            fallback_name = get_first_owner_display_name(source_todo)
            title_only_task = {
                "id": tid,
                "type": TODO_TYPE,
                "promiseDate": None,
                "riskLevel": None,
                "owner_user_guid": fallback_guid,
                "owner_display_name": fallback_name,
                "source_todo_title": source_title,
                "lastModifyDate": source_todo.get("lastModifyDate", ""),
                "days_not_updated": raw_days_not_updated,
                "display_not_updated_count": display_not_updated_count,
                "has_incomplete_owner": has_incomplete_owner,
                "format_errors": all_format_errors,
                "new_title": new_title,
            }

            try:
                result = update_todo(title_only_task)
                if result is not None:
                    update_success_count += 1
            except Exception as e:
                update_fail_count += 1
                print(f"❌ 更新失败(title only) todo={tid} error={e}")
                print(traceback.format_exc())
                fail_records.append({
                    "fail_type": "update_fail",
                    "owner_name": fallback_name,
                    "owner_dept": "",
                    "todo_code": todo_code,
                    "task_title": source_title,
                    "reason": f"标题更新失败: {e}",
                })

print("\n==========================================")
print("🏁 Todo 日报回写应用完成")
print("==========================================\n")

print("================ 执行统计 ================")
print(f"处理项目数: {total_projects}")
print(f"命中日报数: {hit_notes}")
print(f"解析任务数: {total_parsed_tasks}")
print(f"待更新任务数: {total_update_tasks}")
print(f"新增进展成功: {progress_success_count}")
print(f"新增进展失败: {progress_fail_count}")
print(f"字段更新成功(promiseDate/riskLevel/title): {update_success_count}")
print(f"字段更新失败: {update_fail_count}")
print(f"跳过(默认模板): {skip_default_total}")
print(f"跳过(未匹配 Todo): {skip_no_match_count}")
print(f"跳过(无有效变化/缺少owner): {skip_no_change_count}")
print(f"格式错误卡片发送成功: {card_success_count}")
print(f"格式错误卡片发送失败/跳过: {card_fail_count}")
print(f"无格式错误无需发卡片: {card_skip_count}")
print("==========================================\n")

# 报告前兜底清洗：理论上源头已经不会产生 0 次未更新，
# 这里再做一层防御，避免脏数据进入报告展示。
clean_fail_records = []
for r in fail_records:
    if r.get("fail_type") == "not_updated" and int(r.get("days_not_updated") or 0) <= 0:
        continue
    clean_fail_records.append(r)

report_data = {
    "total_parsed_tasks": total_parsed_tasks,
    "total_update_tasks": total_update_tasks,
    "progress_success_count": progress_success_count,
    "update_success_count": update_success_count,
    "success_owners": success_owners,
    "captured_owners": captured_owners,
    "today_update_count": today_update_count,
    "today_update_owners": today_update_owners,
    "fail_records": clean_fail_records,
}

report_result = write_report(TARGET_DATE, report_data)
if report_result:
    doc_guid, report_title = report_result
    send_report_message(report_title, doc_guid, report_data)

if progress_fail_count > 0 or update_fail_count > 0:
    raise Exception(
        f"存在 {progress_fail_count} 个新增进展失败、"
        f"{update_fail_count} 个字段更新失败，请查看日志"
    )
