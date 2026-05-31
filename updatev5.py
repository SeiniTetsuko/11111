# -*- coding: utf-8 -*-
"""
Todo 日报回写应用 v8

核心规则：
1. 解析区标题兼容：
   - TodoList任务进展
   - TodoList 任务进展
   - 平台项目重要任务进展
   - WBS任务进展
   - WBS 任务进展

2. todo_id 提取兼容：
   - TODO#51004
   - WBS#51004
   - 【TODO#51004】
   - 【WBS#51004】
   - 【51004】

3. progress 回写：
   - 每个 owner 写自己的进展。
   - 使用日报 h3 @人名匹配到的 owner.userGuid 调 insertProgress。

4. promiseDate / riskLevel 回写：
   - 按 TodoList owners 顺序顺延。
   - promiseDate 和 riskLevel 分别取第一个有效贡献者。
   - 两个字段贡献者可以不同，贡献者不同时分次 update。

5. title 标签：
   - 只看 TodoList 第一位 owner 的填写状态。
   - 第一 owner 三项均完整且格式正确：清除【X次未更新】/【格式错误】/【回写失败】
   - 第一 owner 任一字段为空：加【X次未更新】
   - 第一 owner 任一字段格式错误：加【格式错误】
   - TodoList 有任务但日报没有该 todo_id：加【回写失败】
   - 未更新和格式错误可同时存在，如【2次未更新】【格式错误】真实标题。

6. 报告统计：
   - 总任务数：TodoList 拉到的任务总数，涉及 owner 总数。
   - 报告不再展示个人级成功更新，只展示未更新详情与异常/缺失记录。
   - 未更新：进展/时间/风险任一未填写的 owner-task 数量，涉及 owner 总数。
   - 格式错误：时间/风险任一格式错误的 owner-task 数量，涉及 owner 总数。
   - 失败：TodoList 有但日报未找到 todo_id 的任务数量，涉及 owner 总数。
   - 任务名称统一 strip 掉【X次未更新】/【格式错误】/【回写失败】标签。
"""

import builtins
import sys
import re
import csv
import json
import time
import traceback
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from collections import OrderedDict

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook

from zenv import get_zdkit_env
from zdbase import ZFile  # 保留平台兼容，不直接依赖

# =============================================================================
# 日志编码兜底
# =============================================================================

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# =============================================================================
# print flush patch
# =============================================================================

if not getattr(builtins.print, "_patched_flush", False):
    _original_print = builtins.print

    def print(*args, **kwargs):
        kwargs.setdefault("flush", True)
        _original_print(*args, **kwargs)

    print._patched_flush = True
    builtins.print = print

# =============================================================================
# 固定 API 路由
# =============================================================================

BASE_URL = "https://workspace.cxmt.com"
TODO_BASE_URL = "http://share.cxmt.com"

TODO_LIST_API = "/openapi/todo/todoList"
TODO_UPDATE_API = "/openapi/todo/update"
TODO_INSERT_PROGRESS_API = "/openapi/todo/insertProgress"
TODO_PROGRESS_LIST_API = "/openapi/todo/progressList"

GET_DOC_API = "/platform/ws/noteInfo/getDocJson"
TOKEN_API = "/api/user/platform/getAccessToken"
DOC_TREE_ROUTE = "/platform/api/main/doc/treeList"
SIGNED_URL_ROUTE = "/platform/api/main/storage/getSignedUrl"

MESSAGE_SEND_ROUTE = "/middle/server/api/msg/send"
WORKSPACE_SAVE_ROUTE = "/platform/api/main/workspace/save"
MD_INSERT_ROUTE = "/middle/server/api/file/md/insert"

# =============================================================================
# 固定业务参数
# =============================================================================

TODO_STATUS_LIST = ["Open", "Ongoing", "Delay"]
TODO_IS_LEAF = False
TODO_TYPE = "wbs"

MESSAGE_TEMPLATE_ID = "80"
PLATFORM_TYPE = "all"

RISK_MAP_CN_TO_EN = {
    "高": "high",
    "中": "middle",
    "低": "low",
}

RISK_MAP_EN_TO_CN = {
    "high": "高",
    "middle": "中",
    "low": "低",
}

WBS_SECTION_TITLES = [
    "TodoList任务进展",
    "TodoList 任务进展",
    "平台项目重要任务进展",
    "WBS任务进展",
    "WBS 任务进展",
]

DEFAULT_PROGRESS_VALUES = {
    "完成了xxx",
    "开发了xxx",
    "请删除本占位符，并填写今日实际工作进展",
    "请删除本占位符，并填写今日实际完成内容",
    "请在此处填写今日实际工作进展",
    "请在此处填写今日实际完成内容",
}

# =============================================================================
# 全局配置加载
# =============================================================================

zenv_obj = get_zdkit_env()
BASE_URL = zenv_obj.zdkit._http_client.config.get("url")

try:
    with open(config_file.path, "r", encoding="utf-8") as config_fp:
        config = json.load(config_fp)
except Exception as e:
    print(f"❌ 配置文件读取失败: {e}")
    raise

AK = config.get("ak")
SK = config.get("sk")
ORG_GUID = config.get("org_guid")
USER_GUID = config.get("user_guid")

MAPPING_GUID = config.get("mapping_guid")
TARGET_DATE = config.get("target_date") or datetime.now().strftime("%Y-%m-%d")

TODO_PROJECT_ID = config.get("todo_project_id")
if isinstance(TODO_PROJECT_ID, str):
    TODO_PROJECT_ID = [TODO_PROJECT_ID]

FORMAT_ERROR_CARD_TITLE = "📍 WBS任务格式填写错误提醒"
FORMAT_ERROR_BUTTON_URL = "http://workshare.cxmt.com/project/list"
FORMAT_ERROR_SENDER_GUID = USER_GUID

if not AK or not SK:
    print("❌ 配置错误：缺少 ak / sk")
    raise ValueError("配置错误：缺少 ak / sk")

if not USER_GUID:
    print("❌ 配置错误：缺少 user_guid")
    raise ValueError("配置错误：缺少 user_guid")

if not MAPPING_GUID:
    print("❌ 配置错误：缺少 mapping_guid")
    raise ValueError("配置错误：缺少 mapping_guid")

if not TODO_PROJECT_ID:
    print("❌ 配置错误：缺少 todo_project_id")
    raise ValueError("配置错误：缺少 todo_project_id")

REPORT_TARGET_PROJECT_GUID = config.get("report_target_project_guid", "")
REPORT_UPDATE_PARENT_GUID = config.get("report_update_parent_guid", "")

REPORT_SENDER_GUID = config.get("report_sender_guid", [])
if isinstance(REPORT_SENDER_GUID, str):
    REPORT_SENDER_GUID = [REPORT_SENDER_GUID]
REPORT_SENDER_GUID = [x for x in REPORT_SENDER_GUID if x]

REPORT_WEBHOOK = config.get("report_webhook", [])
if isinstance(REPORT_WEBHOOK, str):
    REPORT_WEBHOOK = [REPORT_WEBHOOK]
REPORT_WEBHOOK = [x for x in REPORT_WEBHOOK if x]

# =============================================================================
# 通用 HTTP 工具
# =============================================================================

session = requests.Session()

retry_strategy = Retry(
    total=5,
    connect=5,
    read=5,
    backoff_factor=2,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["HEAD", "GET", "POST", "PUT", "DELETE", "OPTIONS"],
    raise_on_status=False,
)

adapter = HTTPAdapter(
    max_retries=retry_strategy,
    pool_connections=10,
    pool_maxsize=10,
)

session.mount("http://", adapter)
session.mount("https://", adapter)


def request_with_retry(method, url, max_retries=3, **kwargs):
    kwargs.setdefault("timeout", 30)
    last_error = None

    for attempt in range(1, max_retries + 1):
        try:
            headers = kwargs.pop("headers", {}) or {}
            headers["Connection"] = "close"
            headers.setdefault("User-Agent", "Mozilla/5.0 PythonRequests")

            if method.lower() == "post":
                response = session.post(url, headers=headers, **kwargs)
            else:
                response = session.get(url, headers=headers, **kwargs)

            return response

        except (
            requests.exceptions.Timeout,
            requests.exceptions.ConnectionError,
            requests.exceptions.ChunkedEncodingError,
        ) as e:
            last_error = e
            wait = min(2 ** attempt, 10)

            print(
                f"⚠️ 请求失败 attempt={attempt}/{max_retries} "
                f"wait={wait}s url={url} error={repr(e)}"
            )

            if attempt < max_retries:
                time.sleep(wait)
            else:
                raise last_error

# =============================================================================
# token / headers
# =============================================================================


def get_access_token():
    print("🔑 获取 AccessToken")

    response = request_with_retry(
        "post",
        BASE_URL + TOKEN_API,
        json={"ak": AK, "sk": SK},
        timeout=30,
    )

    response.raise_for_status()
    response_json = response.json()

    token = (response_json.get("data") or {}).get("accessToken")

    if not token:
        raise Exception(
            f"获取 token 失败: url={BASE_URL + TOKEN_API}, response={response_json}"
        )

    print("✅ AccessToken 获取成功")
    return token


def get_headers(user_guid=None):
    token = get_access_token()

    return {
        "Access-Token": token,
        "ak": AK,
        "X-User-GUID": user_guid or USER_GUID,
        "Content-Type": "application/json",
    }


def get_todo_headers(user_guid=None):
    return {
        "x-user-guid": user_guid or USER_GUID,
        "Content-Type": "application/json",
    }


def get_headers_with_ak(user_guid=None):
    token = get_access_token()
    return {
        "Access-Token": token,
        "ak": AK,
        "X-User-GUID": user_guid or USER_GUID,
        "Content-Type": "application/json",
    }


def send_message_api(receiver_guids, title, content, sender_guid="", interactive_content=None):
    payload = {
        "template_id": MESSAGE_TEMPLATE_ID,
        "receiver_guid": receiver_guids,
        "content": content,
        "org_guid": ORG_GUID,
        "title": title,
        "platform_type": PLATFORM_TYPE,
    }
    if interactive_content is not None:
        payload["interactive_content"] = json.dumps(interactive_content, ensure_ascii=False)

    return request_with_retry(
        "post",
        BASE_URL + MESSAGE_SEND_ROUTE,
        headers=get_headers_with_ak(user_guid=sender_guid),
        json=payload,
        timeout=30,
    )

# =============================================================================
# 基础工具
# =============================================================================


def safe_int(value):
    try:
        return int(value)
    except Exception:
        return None


def is_wbs_section_title(text):
    text = str(text or "").strip()
    return any(title in text for title in WBS_SECTION_TITLES)


def extract_todo_id_from_title_line(title_line):
    """
    兼容：
    - TODO#51004
    - WBS#51004
    - 【TODO#51004】
    - 【WBS#51004】
    - 【51004】

    如果同时存在 TODO#51004 和 subtodo 【30177】，优先取 TODO/WBS 后面的 51004。
    如果只有多个纯数字【51004】&【30177】，兜底取第一个纯数字。
    """
    text = str(title_line or "")

    m = re.search(r"(?:TODO|Todo|todo|WBS|wbs)\s*#\s*(\d+)", text)
    if m:
        return safe_int(m.group(1))

    m = re.search(r"【\s*(?:TODO|Todo|todo|WBS|wbs)\s*#\s*(\d+)\s*】", text)
    if m:
        return safe_int(m.group(1))

    pure_number_matches = re.findall(r"【\s*(\d+)\s*】", text)
    if not pure_number_matches:
        return None

    return safe_int(pure_number_matches[0])


def normalize_date_for_compare(date_value):
    text = str(date_value or "").strip()
    if not text:
        return ""

    normalized = parse_loose_date(text)
    if normalized:
        return normalized

    return text


def parse_loose_date(date_value):
    """
    宽松解析用户填写的日期，并统一转成 YYYY-MM-DD。

    支持：
    - 2026-5-4
    - 2026-05-04
    - 2026/5/4
    - 2026.5.4
    - 20260504

    不支持：
    - 明天
    - 5月4日
    - 2026年5月4日
    """
    text = str(date_value or "").strip()
    if not text:
        return None

    patterns = [
        r"^(\d{4})-(\d{1,2})-(\d{1,2})$",
        r"^(\d{4})/(\d{1,2})/(\d{1,2})$",
        r"^(\d{4})\.(\d{1,2})\.(\d{1,2})$",
    ]

    for pattern in patterns:
        m = re.match(pattern, text)
        if not m:
            continue

        y, mo, d = m.groups()
        try:
            dt = datetime(int(y), int(mo), int(d))
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            return None

    m = re.match(r"^(\d{4})(\d{2})(\d{2})$", text)
    if m:
        y, mo, d = m.groups()
        try:
            dt = datetime(int(y), int(mo), int(d))
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            return None

    return None

# =============================================================================
# mapping 文件读取
# =============================================================================


def get_signed_url(category_guid):
    print(f"🔗 获取签名 URL: category_guid={category_guid}")

    response = request_with_retry(
        "get",
        BASE_URL + SIGNED_URL_ROUTE,
        headers=get_headers(),
        params={"categoryGuid": category_guid},
        timeout=30,
    )

    response.raise_for_status()
    response_json = response.json()
    signed_url = (response_json.get("data") or {}).get("signedUrl")

    if not signed_url:
        raise Exception(
            f"获取签名 URL 失败: category_guid={category_guid}, response={response_json}"
        )

    return signed_url


def parse_xlsx_mapping(content_bytes):
    workbook = load_workbook(
        filename=BytesIO(content_bytes),
        read_only=True,
        data_only=True,
    )

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValueError("xlsx 文件为空")

    headers = [str(x).strip() if x is not None else "" for x in headers]
    rows = []

    for excel_row in rows_iter:
        row_dict = {}

        for idx, value in enumerate(excel_row):
            if idx >= len(headers):
                continue
            row_dict[headers[idx]] = str(value).strip() if value is not None else ""

        rows.append({
            "dept": row_dict.get("Dept", ""),
            "project_guid": row_dict.get("project_guid", ""),
            "work_log_folder_guid": row_dict.get("work_log_folder_guid", ""),
            "project_name": row_dict.get("project_name", ""),
        })

    return rows


def parse_csv_mapping(text):
    stream = StringIO(text)
    reader = csv.DictReader(stream)

    rows = []

    for row in reader:
        rows.append({
            "dept": row.get("Dept", ""),
            "project_guid": row.get("project_guid", ""),
            "work_log_folder_guid": row.get("work_log_folder_guid", ""),
            "project_name": row.get("project_name", ""),
        })

    return rows


def load_mapping_from_guid(mapping_guid):
    print(f"🚀 开始读取 mapping: {mapping_guid}")

    signed_url = get_signed_url(mapping_guid)

    response = request_with_retry(
        "get",
        signed_url,
        timeout=120,
        stream=True,
        headers={"Connection": "close"},
    )

    response.raise_for_status()
    content = response.content
    signed_url_lower = signed_url.lower()

    if ".xlsx" in signed_url_lower:
        rows = parse_xlsx_mapping(content)
    else:
        rows = parse_csv_mapping(content.decode("utf-8-sig", errors="ignore"))

    print(f"✅ mapping 读取完成: 有效行数={len(rows)}")
    return rows

# =============================================================================
# Todo API
# =============================================================================


def fetch_todo_list():
    print(f"🚀 获取 Todo 列表: projectIds={TODO_PROJECT_ID}")

    request_body = {
        "projectIds": TODO_PROJECT_ID,
        "todoStatus": TODO_STATUS_LIST,
        "isLeaf": TODO_IS_LEAF,
        "type": TODO_TYPE,
    }

    response = request_with_retry(
        "post",
        TODO_BASE_URL + TODO_LIST_API,
        headers=get_todo_headers(),
        json=request_body,
        timeout=60,
    )

    response.raise_for_status()
    response_json = response.json()

    if response_json.get("code") not in (0, 200, "0", "200", None):
        print(
            f"⚠️ Todo API 返回 code 异常: "
            f"code={response_json.get('code')}, msg={response_json.get('msg')}"
        )

    todo_list = response_json.get("data") or []
    print(f"✅ 获取 Todo 数量: {len(todo_list)}")

    return todo_list


def build_todo_index(todo_list):
    todo_index = {}
    duplicate_ids = set()

    for todo in todo_list:
        todo_id = todo.get("id")
        if todo_id in (None, ""):
            continue

        todo_id_int = safe_int(todo_id)
        if todo_id_int is None:
            continue

        if todo_id_int in todo_index:
            duplicate_ids.add(todo_id_int)

        todo_index[todo_id_int] = todo

    if duplicate_ids:
        print(f"⚠️ Todo id 存在重复，已以后出现者覆盖: {sorted(duplicate_ids)}")

    print(f"✅ Todo 索引构建完成: {len(todo_index)} 条")
    return todo_index


def build_dept_user_guid_map(todo_list):
    dept_map = OrderedDict()

    for todo in todo_list:
        owners = todo.get("owners", []) or []

        for owner in owners:
            dept = str(owner.get("fullDeptName") or "").strip()
            user_guid = str(owner.get("userGuid") or "").strip()

            if not dept or not user_guid:
                continue

            if dept not in dept_map:
                dept_map[dept] = []

            if user_guid not in dept_map[dept]:
                dept_map[dept].append(user_guid)

    print(f"✅ 部门用户索引构建完成: dept_count={len(dept_map)}")
    return dept_map


def get_first_owner_user_guid(todo):
    owners = todo.get("owners", []) or []
    for owner in owners:
        user_guid = owner.get("userGuid")
        if user_guid:
            return user_guid
    return None


def get_first_owner_display_name(todo):
    owners = todo.get("owners", []) or []
    for owner in owners:
        return owner.get("displayName") or owner.get("name") or owner.get("englishName") or owner.get("loginName") or ""
    return ""


def get_first_owner_identity(source_todo):
    owners = source_todo.get("owners", []) or []
    if not owners:
        return None

    owner = owners[0]

    return {
        "user_guid": str(owner.get("userGuid") or "").strip(),
        "name": str(owner.get("name") or "").strip(),
        "english_name": str(owner.get("englishName") or "").strip(),
        "display_name": (
            owner.get("displayName")
            or owner.get("name")
            or owner.get("englishName")
            or ""
        ),
        "dept": str(owner.get("fullDeptName") or "").strip(),
        "owner_obj": owner,
    }


def find_first_owner_parsed_task(source_todo, group_tasks):
    """
    找到 TodoList 第一位 owner 在日报中的填写记录。
    只用于 title 标签判断，不用于 promiseDate/riskLevel 顺延。
    """
    first_owner = get_first_owner_identity(source_todo)
    if not first_owner:
        return None, None

    first_name = first_owner.get("name", "")
    first_english_name = first_owner.get("english_name", "")

    for pt in group_tasks:
        pt_name = str(pt.get("owner_name") or "").strip()
        if pt_name and (pt_name == first_name or pt_name == first_english_name):
            return pt, first_owner

    return None, first_owner


def match_owner_by_name(todo, owner_name):
    if not owner_name:
        return None

    owners = todo.get("owners", []) or []

    for owner in owners:
        name = str(owner.get("name") or "").strip()
        if name == owner_name:
            return owner

    for owner in owners:
        english_name = str(owner.get("englishName") or "").strip()
        if english_name == owner_name:
            return owner

    return None


def get_todo_code_for_display(todo):
    code = str(todo.get("code") or "").strip()
    if code:
        return code

    todo_id = todo.get("id")
    if todo_id not in (None, ""):
        return f"WBS#{todo_id}"

    return ""


def get_todo_related_project_name(todo, fallback_project_name=""):
    related_project = todo.get("relatedProject") or {}
    return related_project.get("name") or fallback_project_name or "未命名项目"


def get_todo_promise_display(todo):
    """
    早上写入日报时使用：
    将 Todo 接口返回的 promiseDate 转成可展示/可回写解析的日期。
    如果接口没有值，则保留模板占位 YYYY-MM-DD。
    """
    promise = normalize_date_for_compare(todo.get("promiseDate"))
    return promise if promise else "YYYY-MM-DD"


def get_todo_risk_display(todo):
    """
    早上写入日报时使用：
    将 Todo 接口返回的 riskLevel 转成中文 高/中/低。
    如果接口没有值，则保留模板占位 高/中/低。
    """
    risk = str(todo.get("riskLevel") or "").strip()
    return RISK_MAP_EN_TO_CN.get(risk, risk) if risk else "高/中/低"


def build_prefilled_todo_fields_html(todo):
    """
    早上 Todo -> 日报写入应用可直接复用：
    生成预填的承诺完成时间 / 风险等级字段。

    输出示例：
    <p><strong>&lt;承诺完成时间：2026-05-04&gt;:</strong><br></p>
    <p><strong>&lt;风险等级：中&gt;:</strong><br></p>
    """
    promise_display = get_todo_promise_display(todo)
    risk_display = get_todo_risk_display(todo)

    return (
        f"<p><strong>&lt;承诺完成时间：{promise_display}&gt;:</strong><br></p>\n"
        f"<p><strong>&lt;风险等级：{risk_display}&gt;:</strong><br></p>"
    )


def count_unique_todo_owners(todo_list):
    owners = set()

    for todo in todo_list:
        for owner in todo.get("owners", []) or []:
            user_guid = owner.get("userGuid")
            name = owner.get("name") or owner.get("englishName") or owner.get("displayName") or ""

            if user_guid:
                owners.add(user_guid)
            elif name:
                owners.add(name)

    return len(owners)


def count_unique_owners(records):
    owners = set()

    for r in records:
        owner_name = r.get("owner_name", "")
        owner_dept = r.get("owner_dept", "")
        if owner_name or owner_dept:
            owners.add((owner_name, owner_dept))

    return len(owners)


def dedupe_records(records, key_func):
    """
    报告记录去重。
    防止多 mapping 行、同一个 note 多次处理、多 owner 聚合时重复 append。
    """
    result = []
    seen = set()

    for record in records:
        key = key_func(record)
        if key in seen:
            continue
        seen.add(key)
        result.append(record)

    return result


def build_task_unupdated_summary(unupdated_records):
    """
    从个人 owner-task 未更新记录聚合为任务级未更新汇总。
    一个 todo_id 只出现一次，多 owner 合并展示。
    """
    task_map = OrderedDict()

    for record in unupdated_records:
        todo_id = record.get("todo_id") or record.get("todo_code")
        if not todo_id:
            continue

        if todo_id not in task_map:
            task_map[todo_id] = {
                "todo_id": todo_id,
                "todo_code": record.get("todo_code", ""),
                "task_title": strip_title_prefixes(record.get("task_title", "")),
                "days_not_updated": record.get("days_not_updated", ""),
                "owners": OrderedDict(),
            }

        owner_name = record.get("owner_name", "")
        owner_dept = record.get("owner_dept", "")
        owner_key = (owner_name, owner_dept)

        if owner_name or owner_dept:
            owner_display = owner_name
            if owner_dept:
                owner_display = f"{owner_display}({owner_dept})"
            task_map[todo_id]["owners"][owner_key] = owner_display

    return list(task_map.values())


def get_owner_name_candidates(owner):
    """
    Todo owner 与日报 h3 @人名匹配用。
    """
    candidates = set()

    for key in ("name", "englishName", "displayName", "loginName"):
        value = str(owner.get(key) or "").strip()
        if value:
            candidates.add(value)

    return candidates


def parsed_task_matches_owner(parsed_task, owner):
    pt_name = str(parsed_task.get("owner_name") or "").strip()
    if not pt_name:
        return False

    return pt_name in get_owner_name_candidates(owner)


def collect_missing_owner_unupdated_records(source_todo, group_tasks):
    """
    多 owner 任务专用：
    只要 todo_id 已经出现在日报中，就不算任务级缺失；
    但 TodoList owners 中如果有某个 owner 没有在该 todo_id 下填写，则该 owner 计入个人未更新。

    注意：
    - 这里不产生“未在日报中填写该任务”这种异常文案。
    - 这里只产生字段级未更新：缺少进展描述、缺少承诺完成时间、缺少风险等级。
    """
    records = []

    todo_id = safe_int(source_todo.get("id"))
    todo_code = get_todo_code_for_display(source_todo)
    real_title = strip_title_prefixes(source_todo.get("title", ""))
    days = calc_days_since_modify(source_todo.get("lastModifyDate", ""))

    owners = source_todo.get("owners", []) or []

    for owner in owners:
        matched = False

        for parsed_task in group_tasks:
            if parsed_task_matches_owner(parsed_task, owner):
                matched = True
                break

        if matched:
            continue

        owner_name = (
            owner.get("name")
            or owner.get("englishName")
            or owner.get("displayName")
            or owner.get("loginName")
            or ""
        )
        owner_dept = owner.get("fullDeptName", "")

        records.append({
            "todo_id": todo_id,
            "owner_name": owner_name,
            "owner_dept": owner_dept,
            "todo_code": todo_code,
            "task_title": real_title,
            "missing_fields": "进展描述、承诺完成时间、风险等级",
            "days_not_updated": days,
        })

    return records

# =============================================================================
# 日报查找
# =============================================================================


def get_tree_node_title(node):
    for key in ("dataTitle", "title", "name", "fileName", "filename"):
        value = node.get(key)
        if value:
            return str(value).strip()
    return ""


def get_tree_node_guid(node):
    for key in ("categoryGuid", "dataGuid", "guid", "fileGuid", "id"):
        value = node.get(key)
        if value:
            return str(value)
    return ""


def get_date_title_variants(date_str):
    dt = datetime.strptime(date_str, "%Y-%m-%d")

    return [
        dt.strftime("%Y-%m-%d"),
        dt.strftime("%Y/%m/%d"),
        dt.strftime("%Y%m%d"),
        dt.strftime("%Y.%m.%d"),
        dt.strftime("%Y年%m月%d日"),
        f"{dt.year}年{dt.month}月{dt.day}日",
    ]


def get_project_access_user_guids(dept, dept_user_guid_map):
    dept = str(dept or "").strip()

    if not dept:
        return []

    user_guids = dept_user_guid_map.get(dept) or []

    if not user_guids:
        print(f"⚠️ 未找到部门对应用户: dept={dept}")
        return []

    print(f"✅ 匹配部门访问身份: dept={dept}, candidate_count={len(user_guids)}")
    return user_guids


def is_daily_note_title(title, target_date):
    title = title or ""
    return any(x in title for x in get_date_title_variants(target_date))


def list_folder_nodes(project_guid, folder_guid, user_guid=None):
    response = request_with_retry(
        "post",
        BASE_URL + DOC_TREE_ROUTE,
        headers=get_headers(user_guid=user_guid),
        json={
            "projectGuid": project_guid,
            "parentGuid": folder_guid,
        },
        timeout=60,
    )

    response.raise_for_status()
    response_json = response.json()

    return response_json.get("data") or []


def find_today_daily_note(project_guid, folder_guid, target_date, project_name="", user_guid=None):
    print(
        f"🔎 查找日报: project_guid={project_guid}, "
        f"folder_guid={folder_guid}, date={target_date}, project_name={project_name}"
    )

    nodes = list_folder_nodes(project_guid, folder_guid, user_guid=user_guid)

    for node in nodes:
        title = get_tree_node_title(node)
        guid = get_tree_node_guid(node)

        if not title or not guid:
            continue

        if not is_daily_note_title(title, target_date):
            continue

        print(f"✅ 命中日报: {title} | note_guid={guid}")

        return {
            "note_guid": guid,
            "note_title": title,
            "node": node,
        }

    return None


def get_doc_json(doc_id, user_guid=None):
    print(f"📄 获取日报内容: doc_id={doc_id}")

    response = request_with_retry(
        "get",
        BASE_URL + GET_DOC_API,
        headers=get_headers(user_guid=user_guid),
        params={"docId": doc_id},
        timeout=60,
    )

    response.raise_for_status()
    return response.json()


def try_get_daily_note_and_doc(
    project_guid,
    folder_guid,
    target_date,
    project_name,
    candidate_user_guids,
):
    for user_guid in candidate_user_guids:
        print(f"🔐 尝试项目访问身份: user_guid={user_guid}, project={project_name}")

        try:
            daily_note = find_today_daily_note(
                project_guid=project_guid,
                folder_guid=folder_guid,
                target_date=target_date,
                project_name=project_name,
                user_guid=user_guid,
            )

            if not daily_note:
                print(f"ℹ️ 当前用户未找到日报: user_guid={user_guid}")
                continue

            note_guid = daily_note["note_guid"]
            doc_json = get_doc_json(note_guid, user_guid=user_guid)

            print(f"✅ 项目访问身份验证成功: user_guid={user_guid}, note={note_guid}")

            return {
                "daily_note": daily_note,
                "doc_json": doc_json,
                "access_user_guid": user_guid,
            }

        except Exception as e:
            print(f"⚠️ 项目访问失败: user_guid={user_guid}, error={e}")
            continue

    return None

# =============================================================================
# note json 解析
# =============================================================================


def extract_block_text(block):
    texts = []
    _collect_text_from_node(block, texts)
    return "".join(texts).strip()


def _collect_text_from_node(node, texts):
    if isinstance(node, list):
        for item in node:
            _collect_text_from_node(item, texts)
        return

    if not isinstance(node, dict):
        return

    if node.get("type") == "text":
        text = node.get("text", "")
        if text:
            texts.append(text)

    for child in node.get("content", []) or []:
        _collect_text_from_node(child, texts)


def _collect_block_containers(node, result):
    if isinstance(node, list):
        for item in node:
            _collect_block_containers(item, result)
        return

    if not isinstance(node, dict):
        return

    if node.get("type") == "blockContainer":
        result.append(node)
        return

    for child in node.get("content", []) or []:
        _collect_block_containers(child, result)


def _get_block_primary_type(block_container):
    for child in block_container.get("content", []) or []:
        child_type = child.get("type", "")
        if child_type:
            return child_type
    return ""


def _get_heading_level(block_container):
    for child in block_container.get("content", []) or []:
        if child.get("type") == "heading":
            attrs = child.get("attrs", {}) or {}
            level = attrs.get("level", "")
            return str(level)
    return ""


def _collect_owner_name_from_block(block_container):
    texts = []
    mention_labels = []

    def _walk(node):
        if isinstance(node, list):
            for item in node:
                _walk(item)
            return
        if not isinstance(node, dict):
            return
        if node.get("type") == "text":
            t = node.get("text", "").strip()
            if t:
                texts.append(t)
        if node.get("type") == "mention":
            attrs = node.get("attrs") or {}
            label = str(attrs.get("label") or "").strip()
            if label:
                mention_labels.append(label)
        for child in node.get("content", []) or []:
            _walk(child)

    _walk(block_container)

    full_text = " ".join(texts).strip()

    at_match = re.search(r"@(\S+)", full_text)
    if at_match:
        return at_match.group(1)

    if mention_labels:
        return mention_labels[0]

    return None

# =============================================================================
# progress 判定
# =============================================================================


def normalize_progress_line(line):
    line = str(line or "").strip()
    line = re.sub(r"^[\-\•\*]+[\s]*", "", line).strip()
    line = re.sub(r"^\d+[.\)]\s*", "", line).strip()
    return line


def is_meaningful_progress(progress):
    if not progress:
        return False

    lines = []

    for line in progress.splitlines():
        line = normalize_progress_line(line)

        if not line:
            continue

        if line in DEFAULT_PROGRESS_VALUES:
            continue

        lines.append(line)

    return len(lines) > 0


def extract_actual_progress_for_compare(content):
    """
    从进展记录中提取用户实际写入的进展，去掉系统自动追加的提醒。
    """
    text = str(content or "").strip()
    if not text:
        return ""

    m = re.search(r"【进展描述】[:：]\s*(.*)", text, flags=re.S)
    if m:
        text = m.group(1).strip()

    text = re.split(
        r"\n\s*【(?:状态变更提醒|字段一致提醒|格式检查提醒)\b[^】]*】",
        text,
        maxsplit=1,
        flags=re.S,
    )[0].strip()

    return text


def normalize_progress_for_compare(progress):
    """
    进展变化判断专用：
    只比较用户实际进展，忽略项目符号、空行、系统提醒。
    """
    actual = extract_actual_progress_for_compare(progress)
    lines = []

    for line in str(actual or "").splitlines():
        line = normalize_progress_line(line)
        if not line:
            continue
        if line in DEFAULT_PROGRESS_VALUES:
            continue
        lines.append(line)

    return "\n".join(lines).strip()


def get_latest_progress_by_owner(progress_records):
    """
    progressList 返回记录 -> userGuid 到该用户最新实际进展文本的映射。
    """
    latest = {}

    for record in progress_records or []:
        create_by = record.get("createBy") or {}
        user_guid = str(create_by.get("userGuid") or "").strip()
        if not user_guid:
            continue

        date_time = str(record.get("dateTime") or "")
        content = extract_actual_progress_for_compare(record.get("content", ""))

        if user_guid not in latest or date_time >= latest[user_guid].get("dateTime", ""):
            latest[user_guid] = {
                "dateTime": date_time,
                "content": content,
            }

    return {k: v.get("content", "") for k, v in latest.items()}


def has_progress_changed(actual_progress, previous_progress):
    current = normalize_progress_for_compare(actual_progress)
    previous = normalize_progress_for_compare(previous_progress)

    if not current:
        return False

    return current != previous


def build_owner_unupdated_reasons(task, previous_progress, current_promise, current_risk):
    """
    生成 owner-task 的未更新原因。
    报告只展示未更新详情时使用。

    原因包括：
    - 哪一项未填写
    - 哪一项格式错误
    - 哪一项和上次/当前记录一致
    """
    reasons = []

    if not task.get("has_progress"):
        reasons.append("进展描述未填写")
    elif previous_progress and not has_progress_changed(task.get("actual_progress", ""), previous_progress):
        reasons.append("进展描述和上次填写内容一致")

    if task.get("promiseInvalid"):
        reasons.append(f"承诺完成时间格式错误：{task.get('promiseRaw') or '空'}")
    elif not task.get("has_promise"):
        reasons.append("承诺完成时间未填写")
    elif normalize_date_for_compare(task.get("promiseDate")) == normalize_date_for_compare(current_promise):
        reasons.append("承诺完成时间和上次填写内容一致")

    if task.get("riskInvalid"):
        reasons.append(f"风险等级格式错误：{task.get('riskRaw') or '空'}")
    elif not task.get("has_risk"):
        reasons.append("风险等级未填写")
    elif str(task.get("riskLevel") or "").strip() == str(current_risk or "").strip():
        reasons.append("风险等级和上次填写内容一致")

    return reasons


def collect_owner_display_from_todo(source_todo):
    """
    返回任务涉及的所有 owner 展示名。
    """
    owners = []

    for owner in source_todo.get("owners", []) or []:
        owner_name = (
            owner.get("displayName")
            or owner.get("name")
            or owner.get("englishName")
            or owner.get("loginName")
            or ""
        )
        owner_dept = owner.get("fullDeptName", "")

        if owner_name and owner_dept:
            owners.append(f"{owner_name}({owner_dept})")
        elif owner_name:
            owners.append(owner_name)

    return owners


def build_task_level_unupdated_reason(
    *,
    has_group_progress,
    has_group_promise,
    has_group_risk,
    has_group_format_error,
    group_changed,
    progress_changed,
    promise_changed,
    risk_changed,
    enriched_tasks,
    latest_progress_by_owner,
    current_promise,
    current_risk,
):
    """
    任务级未更新原因：
    当一个任务下所有 owner 合并后，仍无法形成有效更新时使用。

    有效更新定义：
    1. 进展 / 承诺完成时间 / 风险等级 三项均存在；
    2. 且至少一项和上次记录不一致；
    3. 且不存在格式错误。
    """
    reasons = []

    if not has_group_progress:
        reasons.append("进展描述未填写")
    if not has_group_promise:
        reasons.append("承诺完成时间未填写")
    if not has_group_risk:
        reasons.append("风险等级未填写")

    # 格式错误聚合
    format_error_parts = []
    for task in enriched_tasks:
        owner = task.get("progress_owner_display") or task.get("owner_display_name") or ""
        if task.get("promiseInvalid"):
            format_error_parts.append(f"{owner}承诺完成时间格式错误：{task.get('promiseRaw') or '空'}")
        if task.get("riskInvalid"):
            format_error_parts.append(f"{owner}风险等级格式错误：{task.get('riskRaw') or '空'}")

    if format_error_parts:
        reasons.extend(format_error_parts)

    # 三项都存在且无格式错误，但没有任何变化
    if has_group_progress and has_group_promise and has_group_risk and not has_group_format_error and not group_changed:
        same_parts = []

        if not progress_changed:
            # 只列实际写了进展但和上次相同的人
            same_progress_owners = []
            for task in enriched_tasks:
                actual_progress = task.get("actual_progress", "")
                if not is_meaningful_progress(actual_progress):
                    continue

                previous_progress = latest_progress_by_owner.get(task.get("progress_user_guid"), "")
                if previous_progress and not has_progress_changed(actual_progress, previous_progress):
                    owner = task.get("progress_owner_display") or task.get("owner_display_name") or ""
                    if owner:
                        same_progress_owners.append(owner)

            if same_progress_owners:
                same_parts.append("进展描述和上次填写内容一致：" + "、".join(sorted(set(same_progress_owners))))
            else:
                same_parts.append("进展描述和上次填写内容一致")

        if not promise_changed:
            same_parts.append("承诺完成时间和上次填写内容一致")

        if not risk_changed:
            same_parts.append("风险等级和上次填写内容一致")

        reasons.extend(same_parts)

    if not reasons:
        reasons.append("未形成有效更新")

    return "；".join(reasons)

# =============================================================================
# 解析 TodoList任务进展
# =============================================================================


def parse_tasks_from_doc(doc_json):
    content = (doc_json.get("data") or {}).get("content", [])

    all_blocks = []
    _collect_block_containers(content, all_blocks)

    tasks = []
    skip_default_count = 0

    print("🚀 开始解析日报内容（JSON 结构）")

    current_task = None
    wbs_section_started = False
    current_owner_name = None

    for block in all_blocks:
        block_text = extract_block_text(block)
        primary_type = _get_block_primary_type(block)

        if primary_type == "heading":
            level = _get_heading_level(block)

            if level == "2":
                if current_task:
                    skip_default_count = _finalize_task(current_task, tasks, skip_default_count)
                    current_task = None

                if is_wbs_section_title(block_text):
                    if not wbs_section_started:
                        wbs_section_started = True
                        print(f"✅ 检测到 WBS/TodoList 任务区域开始: {block_text}")
                    current_owner_name = None
                else:
                    if wbs_section_started:
                        wbs_section_started = False
                        current_owner_name = None
                        print("ℹ️ WBS/TodoList 任务区域结束")

                continue

            if level == "3":
                if current_task:
                    skip_default_count = _finalize_task(current_task, tasks, skip_default_count)
                    current_task = None

                owner_name = _collect_owner_name_from_block(block)
                if owner_name:
                    current_owner_name = owner_name
                    print(f"👤 检测到 owner 区域: @{current_owner_name}")
                else:
                    current_owner_name = None

                continue

        if not wbs_section_started:
            continue

        if "🚀" in block_text:
            if current_task:
                skip_default_count = _finalize_task(current_task, tasks, skip_default_count)

            todo_id = extract_todo_id_from_title_line(block_text)

            current_task = {
                "todo_id": todo_id,
                "title_line": block_text,
                "owner_name": current_owner_name,

                "promise_raw": None,
                "promise_date": None,
                "promise_invalid": False,

                "risk_raw": None,
                "risk_level": None,
                "risk_invalid": False,

                "progress_lines": [],
                "in_progress_section": False,
            }

            continue

        if not current_task:
            continue

        due_match = re.match(r"<任务截止时间[^>]*>:[ \t]*(.*)", block_text)
        if due_match:
            current_task["in_progress_section"] = False
            continue

        # 承诺完成时间：
        # 兼容两种写法：
        # 1) <承诺完成时间：YYYY-MM-DD>: 2026-05-04
        # 2) <承诺完成时间：2026-05-04>:
        promise_match = re.match(r"<承诺完成时间[：:]?\s*([^>]*)>[:：]?[ \t]*(.*)", block_text)
        if promise_match:
            inside_value = promise_match.group(1).strip()
            after_colon_value = promise_match.group(2).strip()

            # 用户如果在冒号后填写，以冒号后的值为准；
            # 否则读取尖括号内的预填值。
            user_value = after_colon_value or inside_value

            # 模板占位不算有效填写
            if user_value in ("YYYY-MM-DD", "yyyy-mm-dd", "日期"):
                user_value = ""

            current_task["promise_raw"] = user_value

            normalized_date = parse_loose_date(user_value)
            if normalized_date:
                current_task["promise_date"] = normalized_date
            elif user_value:
                current_task["promise_invalid"] = True

            current_task["in_progress_section"] = False
            continue

        # 风险等级：
        # 兼容两种写法：
        # 1) <风险等级：高/中/低>: 中
        # 2) <风险等级：中>:
        risk_match = re.match(r"<风险等级[：:]?\s*([^>]*)>[:：]?[ \t]*(.*)", block_text)
        if risk_match:
            inside_value = risk_match.group(1).strip()
            after_colon_value = risk_match.group(2).strip()

            # 用户如果在冒号后填写，以冒号后的值为准；
            # 否则读取尖括号内的预填值。
            user_value = after_colon_value or inside_value

            # 模板占位不算有效填写
            if user_value in ("高/中/低", "高中低", "风险等级"):
                user_value = ""

            current_task["risk_raw"] = user_value

            if user_value in RISK_MAP_CN_TO_EN:
                current_task["risk_level"] = RISK_MAP_CN_TO_EN[user_value]
            elif user_value:
                current_task["risk_invalid"] = True

            current_task["in_progress_section"] = False
            continue

        progress_match = re.match(r"<今日进展描述[^>]*>[:：]?[ \t]*(.*)", block_text)
        if progress_match:
            current_task["in_progress_section"] = True
            inline_text = progress_match.group(1).strip()

            if inline_text:
                line = normalize_progress_line(inline_text)
                if line:
                    current_task["progress_lines"].append(line)

            continue

        if current_task["in_progress_section"]:
            line = normalize_progress_line(block_text)
            if line:
                current_task["progress_lines"].append(line)

    if current_task:
        skip_default_count = _finalize_task(current_task, tasks, skip_default_count)

    print(f"✅ 解析完成: 有效任务={len(tasks)}, 跳过默认模板={skip_default_count}")
    return tasks, skip_default_count


def _finalize_task(current_task, tasks, skip_default_count):
    if not current_task.get("todo_id"):
        print(f"⚠️ 跳过无法解析 todo_id 的任务行: {current_task.get('title_line', '')}")
        return skip_default_count

    raw_progress = "\n".join(current_task["progress_lines"]).strip()
    has_meaningful_progress = is_meaningful_progress(raw_progress)

    if not has_meaningful_progress:
        raw_progress = ""

    warnings = []

    if current_task["promise_invalid"]:
        warnings.append("- 承诺完成时间填写格式有误，正确格式：YYYY-MM-DD")

    if current_task["risk_invalid"]:
        warnings.append("- 风险等级填写格式有误，正确格式：高/中/低")

    progress_parts = []

    if warnings:
        progress_parts.append(f"【格式检查提醒 {TARGET_DATE}】")
        progress_parts.extend(warnings)

    if has_meaningful_progress:
        if progress_parts:
            progress_parts.append("")
        progress_parts.append(raw_progress)

    progress = "\n".join(progress_parts).strip()

    if not has_meaningful_progress and not warnings:
        print(f"ℹ️ 默认模板内容，仅检查 title 前缀: todo={current_task['todo_id']}")
        skip_default_count += 1

    tasks.append({
        "id": current_task["todo_id"],
        "owner_name": current_task.get("owner_name"),

        "promiseDate": current_task["promise_date"],
        "promiseRaw": current_task["promise_raw"],
        "promiseInvalid": current_task["promise_invalid"],

        "riskLevel": current_task["risk_level"],
        "riskRaw": current_task["risk_raw"],
        "riskInvalid": current_task["risk_invalid"],

        "progress": progress,
        "actual_progress": raw_progress,
        "has_progress": has_meaningful_progress,
        "has_promise": current_task["promise_date"] is not None,
        "has_risk": current_task["risk_level"] is not None,
    })

    return skip_default_count

# =============================================================================
# 状态分类与 title 标签
# =============================================================================

INVISIBLE_CHARS_RE = re.compile(r"[\u200b\u200c\u200d\u2060\ufeff\xa0]")


def normalize_report_text(text):
    """
    报告和 title 清洗专用：
    去掉零宽字符、BOM、不可见空格，避免类似【​1次未更新】无法被正则识别。
    """
    return INVISIBLE_CHARS_RE.sub("", str(text or "")).strip()


_TITLE_PREFIX_RE = re.compile(
    r"^(?:"
    # 兼容：【1次未更新】/【10日未更新】/【1 未更新】
    r"【\s*\d+\s*(?:次|日)?\s*未更新\s*】"
    r"|【\s*格式错误\s*】"
    r"|【\s*回写失败\s*】"
    r")+"
)


def strip_title_prefixes(title):
    """
    去掉 title 前面的系统状态标签，只保留真实任务名。

    兼容：
    - 【1次未更新】测试
    - 【10日未更新】测试
    - 【​1次未更新】测试
    - 【格式错误】测试
    - 【回写失败】测试
    - 【1次未更新】【格式错误】测试
    """
    title = normalize_report_text(title)

    while True:
        new_title = _TITLE_PREFIX_RE.sub("", title).strip()
        if new_title == title:
            break
        title = new_title

    return title

def calc_days_since_modify(last_modify_date_str, now=None):
    if not last_modify_date_str:
        return 0

    try:
        last_dt = datetime.strptime(
            str(last_modify_date_str).strip()[:19], "%Y-%m-%d %H:%M:%S"
        )
    except Exception:
        return 0

    ref = now or datetime.now()

    workdays = 0
    check_date = last_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = ref.replace(hour=0, minute=0, second=0, microsecond=0)
    one_day = timedelta(days=1)

    check_date += one_day

    while check_date <= end_date:
        if check_date.weekday() < 5:
            workdays += 1
        check_date += one_day

    return workdays


def classify_owner_task_status(task):
    """
    单个 owner-task 的状态判断。
    用于：
    1. 第一 owner 判断 title 标签
    2. 所有 owner 做报告统计
    """
    missing_fields = []
    error_fields = []

    if not task.get("has_progress"):
        missing_fields.append("进展描述")

    if not task.get("has_promise"):
        missing_fields.append("承诺完成时间")

    if not task.get("has_risk"):
        missing_fields.append("风险等级")

    if task.get("promiseInvalid"):
        error_fields.append({
            "field": "承诺完成时间",
            "raw": task.get("promiseRaw") or "空",
            "expected": "YYYY-MM-DD",
        })

    if task.get("riskInvalid"):
        error_fields.append({
            "field": "风险等级",
            "raw": task.get("riskRaw") or "空",
            "expected": "高/中/低",
        })

    is_unupdated = len(missing_fields) > 0
    is_format_error = len(error_fields) > 0

    is_success_update = (
        not is_unupdated
        and not is_format_error
        and task.get("has_progress")
        and task.get("has_promise")
        and task.get("has_risk")
    )

    return {
        "is_success_update": is_success_update,
        "is_unupdated": is_unupdated,
        "is_format_error": is_format_error,
        "missing_fields": missing_fields,
        "error_fields": error_fields,
    }


def build_update_title_v2(source_title, last_modify_date, first_owner_status, is_writeback_failed=False):
    """
    title 标签只看第一位 owner。
    回写失败是任务级失败，日报完全没找到该 todo_id 时加。
    """
    current_title = str(source_title or "").strip()
    base_title = strip_title_prefixes(current_title)

    prefixes = []

    if is_writeback_failed:
        prefixes.append("【回写失败】")
    else:
        if first_owner_status.get("is_unupdated"):
            days = calc_days_since_modify(last_modify_date)
            prefixes.append(f"【{max(days, 1)}次未更新】")

        if first_owner_status.get("is_format_error"):
            prefixes.append("【格式错误】")

    new_title = "".join(prefixes) + base_title

    if new_title == current_title:
        return None

    return new_title

# =============================================================================
# enrich
# =============================================================================


def enrich_task_with_todo_context(parsed_task, todo_index, note_project_name=""):
    todo_id = parsed_task["id"]
    source_todo = todo_index.get(todo_id)

    if not source_todo:
        print(
            f"⚠️ 日报中解析到 todo_id={todo_id}，"
            f"但未在 config.todo_project_id 拉取的 Todo 列表中找到，跳过更新"
        )
        return None

    owner_user_guid = get_first_owner_user_guid(source_todo)
    owner_display_name = get_first_owner_display_name(source_todo)

    matched_owner = match_owner_by_name(source_todo, parsed_task.get("owner_name"))
    if matched_owner:
        progress_user_guid = matched_owner.get("userGuid")
        progress_owner_display = (
            matched_owner.get("displayName")
            or matched_owner.get("name")
            or matched_owner.get("englishName")
            or ""
        )
    else:
        progress_user_guid = None
        progress_owner_display = ""

    source_project_name = get_todo_related_project_name(source_todo, fallback_project_name=note_project_name)
    related_project = source_todo.get("relatedProject") or {}
    project_id = related_project.get("id", "")
    todo_code = get_todo_code_for_display(source_todo)

    if not owner_user_guid:
        print(
            f"⚠️ Todo 匹配成功但缺少 owners.userGuid，跳过更新: "
            f"todo_id={todo_id}, title={source_todo.get('title', '')}"
        )
        return None

    if not progress_user_guid:
        owner_name = parsed_task.get("owner_name") or ""
        print(
            f"⚠️ 日报 @人名「{owner_name}」未匹配到 Todo owner，"
            f"进展将使用第一个 owner 身份回写: todo_id={todo_id}"
        )
        progress_user_guid = owner_user_guid
        progress_owner_display = owner_display_name

    current_promise = normalize_date_for_compare(source_todo.get("promiseDate"))
    current_risk = str(source_todo.get("riskLevel") or "").strip()

    update_promise = parsed_task.get("promiseDate")
    update_risk = parsed_task.get("riskLevel")

    unchanged_messages = []

    if update_promise is not None:
        if normalize_date_for_compare(update_promise) == current_promise:
            unchanged_messages.append("- 承诺完成时间和上次更新一致")
            update_promise = None

    if update_risk is not None:
        if update_risk == current_risk:
            unchanged_messages.append("- 风险等级和上次更新一致")
            update_risk = None

    format_errors = []

    if parsed_task.get("promiseInvalid"):
        format_errors.append({
            "field": "承诺完成时间",
            "raw": parsed_task.get("promiseRaw") or "空",
            "expected": "YYYY-MM-DD",
        })

    if parsed_task.get("riskInvalid"):
        format_errors.append({
            "field": "风险等级",
            "raw": parsed_task.get("riskRaw") or "空",
            "expected": "高/中/低",
        })

    progress_content_parts = []

    progress = parsed_task.get("progress", "").strip()
    if progress:
        progress_content_parts.append(f"【进展描述】：\n{progress}")

    change_messages = []

    if update_risk is not None:
        risk_display_new = RISK_MAP_EN_TO_CN.get(update_risk, update_risk)
        risk_display_old = RISK_MAP_EN_TO_CN.get(current_risk, current_risk) if current_risk else "无"
        change_messages.append(f"风险变更为【{risk_display_new}】（原值：{risk_display_old}）")

    if update_promise is not None:
        promise_display_old = current_promise or "无"
        change_messages.append(f"承诺完成时间变更为【{update_promise}】（原值：{promise_display_old}）")

    if change_messages:
        if progress_content_parts:
            progress_content_parts.append("")
        progress_content_parts.append(f"【状态变更提醒 {TARGET_DATE}】")
        progress_content_parts.extend([f"- {m}" for m in change_messages])

    if unchanged_messages:
        if progress_content_parts:
            progress_content_parts.append("")
        progress_content_parts.append(f"【字段一致提醒 {TARGET_DATE}】")
        progress_content_parts.extend(unchanged_messages)

    progress_content = "\n".join(progress_content_parts).strip()

    return {
        "id": todo_id,
        "type": TODO_TYPE,
        "progress_content": progress_content,

        # 注意：这里的 promiseDate/riskLevel 只用于该 owner 自己的进展文案。
        # 真正字段回写会在主流程中按 owner 顺序顺延重新计算。
        "promiseDate": update_promise,
        "riskLevel": update_risk,

        "owner_user_guid": owner_user_guid,
        "owner_display_name": owner_display_name,
        "progress_user_guid": progress_user_guid,
        "progress_owner_display": progress_owner_display,
        "source_todo_title": source_todo.get("title", ""),
        "source_project_name": source_project_name,
        "todo_code": todo_code,
        "format_errors": format_errors,
        "currentPromiseDate": current_promise,
        "currentRiskLevel": current_risk,
        "project_id": project_id,
        "lastModifyDate": source_todo.get("lastModifyDate", ""),

        "actual_progress": parsed_task.get("actual_progress", ""),
        "has_progress": parsed_task.get("has_progress", False),
        "has_promise": parsed_task.get("has_promise", False),
        "has_risk": parsed_task.get("has_risk", False),
        "promiseInvalid": parsed_task.get("promiseInvalid", False),
        "promiseRaw": parsed_task.get("promiseRaw"),
        "riskInvalid": parsed_task.get("riskInvalid", False),
        "riskRaw": parsed_task.get("riskRaw"),
    }

# =============================================================================
# 飞书格式错误提醒卡片
# =============================================================================


def build_format_error_card_content(task):
    project_name = task.get("source_project_name", "未命名项目")
    todo_name = strip_title_prefixes(task.get("source_todo_title", "未命名任务"))
    todo_code = task.get("todo_code", "")
    project_id = task.get("project_id", "")

    task_line = f"【{project_name}】【{todo_name}】【{todo_code}】"

    lines = [
        f"您的任务 **{task_line}** 格式填写有误",
        "",
        "| 项目ID | 项目名称 | 任务Code | 任务名称 |",
        "| --- | --- | --- | --- |",
        f"| {project_id} | {project_name} | {todo_code} | {todo_name} |",
        "",
        "**错误详情：**",
    ]

    for error in task.get("format_errors", []) or []:
        field = error.get("field", "")
        raw_value = error.get("raw", "")
        expected = error.get("expected", "")
        lines.append(f"- **<{field}>** 填写为：**{raw_value}**，应填写格式：**{expected}**")

    lines.extend(["", "请及时前往项目中心更新。"])

    return "\n".join(lines)


def build_format_error_plain_text(task):
    project_name = task.get("source_project_name", "未命名项目")
    todo_name = strip_title_prefixes(task.get("source_todo_title", "未命名任务"))
    todo_code = task.get("todo_code", "")

    error_parts = []
    for error in task.get("format_errors", []) or []:
        field = error.get("field", "")
        raw_value = error.get("raw", "")
        expected = error.get("expected", "")
        error_parts.append(f"<{field}>填写为：{raw_value}，应填写格式为：{expected}")

    error_text = "；".join(error_parts)

    return (
        f"您的任务：【{project_name}】【{todo_name}】【{todo_code}】格式填写有误。"
        f"{error_text}。请前往项目中心及时更新：{FORMAT_ERROR_BUTTON_URL}"
    )


def build_format_error_feishu_card(task):
    card_content = build_format_error_card_content(task)

    return {
        "schema": "2.0",
        "header": {
            "padding": "12px 8px 12px 8px",
            "template": "red",
            "title": {
                "content": FORMAT_ERROR_CARD_TITLE,
                "tag": "plain_text",
            },
        },
        "body": {
            "vertical_spacing": "12px",
            "elements": [
                {
                    "tag": "markdown",
                    "content": card_content,
                    "margin": "0px",
                    "text_size": "normal",
                },
                {
                    "tag": "column_set",
                    "flex_mode": "stretch",
                    "horizontal_spacing": "8px",
                    "margin": "0px",
                    "columns": [
                        {
                            "tag": "column",
                            "width": "weighted",
                            "weight": 1,
                            "elements": [
                                {
                                    "tag": "button",
                                    "type": "primary_filled",
                                    "width": "fill",
                                    "margin": "4px 0px 4px 0px",
                                    "text": {
                                        "tag": "plain_text",
                                        "content": "前往项目中心",
                                    },
                                    "behaviors": [
                                        {
                                            "type": "open_url",
                                            "default_url": FORMAT_ERROR_BUTTON_URL,
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                },
            ],
        },
    }


def send_format_error_card(task):
    format_errors = task.get("format_errors", []) or []
    if not format_errors:
        return False

    receiver_guid = task.get("progress_user_guid")
    if not receiver_guid:
        print(f"⚠️ 格式错误卡片未发送：缺少 receiver progress_user_guid, todo={task.get('id')}")
        return False

    card = build_format_error_feishu_card(task)
    plain_text = build_format_error_plain_text(task)

    print(f"📩 准备发送格式错误提醒卡片: todo={task.get('id')}, receiver={receiver_guid}")

    try:
        response = send_message_api(
            receiver_guids=[receiver_guid],
            title=FORMAT_ERROR_CARD_TITLE,
            content=plain_text,
            sender_guid=FORMAT_ERROR_SENDER_GUID,
            interactive_content=card,
        )

        if hasattr(response, "status_code"):
            ok = response.status_code == 200
            try:
                response_json = response.json()
            except Exception:
                response_json = {}

            if ok and (response_json.get("data") or response_json.get("code") in (0, 200, "0", "200", None)):
                print(f"✅ 格式错误提醒卡片发送成功: todo={task.get('id')}")
                return True

            print(f"❌ 格式错误提醒卡片发送失败: status={response.status_code}, text={getattr(response, 'text', '')}")
            return False

        print(f"✅ 格式错误提醒卡片发送完成: todo={task.get('id')}, response={response}")
        return True

    except Exception as e:
        print(f"❌ 发送格式错误提醒卡片异常: todo={task.get('id')}, error={e}")
        print(traceback.format_exc())
        return False

# =============================================================================
# 报告生成
# =============================================================================


def escape_html(text):
    text = str(text or "")
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def escape_md_cell(text):
    """
    Markdown 表格单元格转义。
    避免 | 和换行破坏表格。
    """
    text = normalize_report_text(text)
    text = text.replace("|", "｜")
    text = text.replace("\r", " ").replace("\n", " ")
    return text.strip()


def create_workspace_doc(project_guid, parent_id, title, tag_list=None):
    print(f"📄 创建文档: title={title}, project_guid={project_guid}, parent_id={parent_id}")

    payload = {
        "categoryName": title,
        "projectGuid": project_guid,
        "parentId": parent_id,
        "categoryType": 2,
        "tagList": tag_list or []
    }

    response = request_with_retry(
        "post",
        BASE_URL + WORKSPACE_SAVE_ROUTE,
        headers=get_headers(),
        json=payload,
        timeout=30
    )

    if response.status_code != 200:
        raise Exception(f"创建文档失败: status={response.status_code}, text={response.text[:500]}")

    response_json = response.json()
    doc_guid = response_json.get("data")

    if not doc_guid:
        raise Exception(f"创建文档返回无效 data: response={response_json}")

    print(f"✅ 文档创建成功: title={title}, doc_guid={doc_guid}")
    return doc_guid


def insert_markdown_to_note(note_guid, markdown_content, writer_user_guid=None):
    if not markdown_content.strip():
        print(f"ℹ️ note={note_guid} 无可写入内容，跳过")
        return None

    print(f"📝 写入报告: note_guid={note_guid}, writer_user_guid={writer_user_guid or USER_GUID}")

    payload = {
        "note_guid": note_guid,
        "markdown_content": markdown_content,
        "mode": "a",
        "location": 1
    }

    response = request_with_retry(
        "post",
        BASE_URL + MD_INSERT_ROUTE,
        headers=get_headers(user_guid=writer_user_guid),
        json=payload,
        timeout=60
    )

    if response.status_code != 200:
        raise Exception(
            f"写入笔记失败: note_guid={note_guid}, "
            f"status={response.status_code}, text={response.text[:500]}"
        )

    response_json = response.json()
    print(f"✅ 写入成功: note_guid={note_guid}")

    return response_json


def build_report_markdown(target_date, report_data):
    total_tasks = report_data.get("total_todo_tasks", 0)
    total_owners = report_data.get("total_todo_owners", 0)

    unupdated_records = report_data.get("unupdated_records", [])
    failed_records = report_data.get("writeback_failed_records", [])

    lines = []

    lines.append("## 任务未更新情况汇总")
    lines.append(f"**总任务数**：{total_tasks}，**涉及 owner 总数**：{total_owners}")
    lines.append(f"**未形成有效更新任务**：{len(unupdated_records)} 个")
    lines.append(f"**异常/缺失记录**：{len(failed_records)} 条")
    lines.append("")
    lines.append("说明：有效更新指任务下所有 owner 的填写合并后，进展描述、承诺完成时间、风险等级均存在，且至少一项和上次更新记录不一致。")
    lines.append("")

    if unupdated_records:
        lines.append("## 未更新任务详情")
        lines.append("")
        lines.append("| 任务Code | 任务名称 | 未更新次数 | 涉及 owner | 未更新情况说明 |")
        lines.append("| --- | --- | --- | --- | --- |")

        for r in unupdated_records:
            code = escape_md_cell(r.get("todo_code", ""))
            title = escape_md_cell(strip_title_prefixes(r.get("task_title", "")))
            days = escape_md_cell(str(r.get("days_not_updated", "")))
            owners = escape_md_cell(r.get("owner_name", ""))
            reason = escape_md_cell(r.get("missing_fields", ""))
            lines.append(f"| {code} | {title} | {days}次 | {owners} | {reason} |")

        lines.append("")

    if failed_records:
        lines.append("## 异常/缺失记录")
        lines.append("说明：这里仅统计任务级缺失或接口异常。")
        lines.append("")
        lines.append("| 负责人 | 部门 | 任务Code | 任务名称 | 失败原因 |")
        lines.append("| --- | --- | --- | --- | --- |")

        for r in failed_records:
            owner = escape_md_cell(r.get("owner_name", ""))
            dept = escape_md_cell(r.get("owner_dept", ""))
            code = escape_md_cell(r.get("todo_code", ""))
            title = escape_md_cell(strip_title_prefixes(r.get("task_title", "")))
            reason = escape_md_cell(r.get("reason", ""))
            lines.append(f"| {owner} | {dept} | {code} | {title} | {reason} |")

        lines.append("")

    return "\n".join(lines)


def build_report_card_content(report_data):
    total_tasks = report_data.get("total_todo_tasks", 0)
    total_owners = report_data.get("total_todo_owners", 0)

    unupdated_records = report_data.get("unupdated_records", [])
    failed_records = report_data.get("writeback_failed_records", [])

    lines = []
    lines.append(f"**总任务数**：{total_tasks}，**涉及 owner 总数**：{total_owners}")
    lines.append(f"**未形成有效更新任务**：{len(unupdated_records)} 个")
    lines.append(f"**异常/缺失记录**：{len(failed_records)} 条")

    if unupdated_records:
        task_names = [
            f"{r.get('todo_code', '')} {strip_title_prefixes(r.get('task_title', ''))}"
            for r in unupdated_records[:10]
        ]
        lines.append("")
        lines.append("**未更新任务**：" + "、".join(task_names))

    return "\n".join(lines)


def build_report_feishu_card(title, card_content, note_url):
    return {
        "schema": "2.0",
        "header": {
            "padding": "12px 8px 12px 8px",
            "template": "blue",
            "title": {
                "content": title,
                "tag": "plain_text"
            }
        },
        "body": {
            "vertical_spacing": "12px",
            "elements": [
                {
                    "tag": "markdown",
                    "content": card_content,
                    "margin": "0px",
                    "text_size": "normal"
                },
                {
                    "tag": "column_set",
                    "flex_mode": "stretch",
                    "horizontal_spacing": "8px",
                    "margin": "0px",
                    "columns": [
                        {
                            "tag": "column",
                            "width": "weighted",
                            "weight": 1,
                            "elements": [
                                {
                                    "tag": "button",
                                    "type": "primary_filled",
                                    "width": "fill",
                                    "margin": "4px 0px 4px 0px",
                                    "text": {
                                        "tag": "plain_text",
                                        "content": "查看报告"
                                    },
                                    "behaviors": [
                                        {
                                            "type": "open_url",
                                            "default_url": note_url
                                        }
                                    ]
                                }
                            ]
                        }
                    ]
                }
            ]
        }
    }


def send_webhook_card(webhook_url, card, max_retries=3, retry_interval=5):
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.post(
                url=webhook_url,
                headers={"Content-Type": "application/json"},
                json={"msg_type": "interactive", "card": card},
                timeout=10
            )
            result = response.json()

            if result.get("code") == 0 or result.get("StatusCode") == 0:
                return result

            if attempt < max_retries:
                print(f"  -> ⚠️ Webhook 发送失败 (尝试 {attempt}/{max_retries}): {result}，{retry_interval}秒后重试...")
                time.sleep(retry_interval)
            else:
                return result
        except Exception as e:
            if attempt < max_retries:
                print(f"  -> ⚠️ Webhook 发送异常 (尝试 {attempt}/{max_retries}): {e}，{retry_interval}秒后重试...")
                time.sleep(retry_interval)
            else:
                raise

    return {"code": -1, "msg": "max retries exceeded"}


def send_report_message(report_title, doc_guid, report_data):
    if not REPORT_WEBHOOK and not REPORT_SENDER_GUID:
        print("ℹ️ 未配置 report_webhook 或 report_sender_guid，跳过消息发送")
        return

    note_url = f"{BASE_URL}/workspace/{doc_guid}"
    card_content = build_report_card_content(report_data)
    card = build_report_feishu_card(report_title, card_content, note_url)

    has_sent_any = False

    if REPORT_WEBHOOK:
        for idx, url in enumerate(REPORT_WEBHOOK, 1):
            try:
                print(f"📢 正在发送群消息 (Webhook {idx}/{len(REPORT_WEBHOOK)})...")
                result = send_webhook_card(url, card)

                if result.get("code") == 0 or result.get("StatusCode") == 0:
                    print(f"  -> ✅ 群消息发送成功: {url[:30]}...")
                    has_sent_any = True
                else:
                    print(f"  -> ❌ 群消息发送失败: {result}")
            except Exception as e:
                print(f"  -> ❌ 群消息发送异常: {e}")

    if REPORT_SENDER_GUID:
        try:
            print(f"📩 正在发送个人消息给 {len(REPORT_SENDER_GUID)} 人...")
            text_content = f"【{report_title}】已生成，请点击查看。\n<a href='{note_url}'>点击查看详情</a>"
            response = send_message_api(
                receiver_guids=REPORT_SENDER_GUID,
                title=report_title,
                content=text_content,
                interactive_content=card,
            )
            if hasattr(response, "status_code") and response.status_code == 200:
                try:
                    resp_json = response.json()
                except Exception:
                    resp_json = {}
                if resp_json.get("data") or resp_json.get("code") in (0, 200, "0", "200", None):
                    print("  -> ✅ 个人消息发送成功")
                    has_sent_any = True
                else:
                    print(f"  -> ❌ 个人消息发送失败: {response.text if hasattr(response, 'text') else '无响应'}")
            else:
                print(f"  -> ❌ 个人消息发送失败: {response.text if response and hasattr(response, 'text') else '无响应'}")
        except Exception as e:
            print(f"  -> ❌ 个人消息发送异常: {e}")

    if has_sent_any:
        print("✅ 报告消息发送完成")
    else:
        print("⚠️ 报告消息未能成功发送到任何目标")


def write_report(target_date, report_data):
    if not REPORT_TARGET_PROJECT_GUID or not REPORT_UPDATE_PARENT_GUID:
        print("ℹ️ 未配置 report_target_project_guid 或 report_update_parent_guid，跳过报告生成")
        return None

    report_title = f"{target_date} 项目管理任务进展更新情况"

    print(f"\n🚀 开始生成任务更新报告: {report_title}")

    doc_guid = create_workspace_doc(
        project_guid=REPORT_TARGET_PROJECT_GUID,
        parent_id=REPORT_UPDATE_PARENT_GUID,
        title=report_title
    )

    markdown_content = build_report_markdown(target_date, report_data)

    print("\n================ 报告内容预览 ================\n")
    print(markdown_content)
    print("\n==============================================\n")

    insert_markdown_to_note(
        note_guid=doc_guid,
        markdown_content=markdown_content
    )

    print(f"✅ 任务更新报告写入完成: {report_title}")

    return doc_guid, report_title

# =============================================================================
# Todo progressList
# =============================================================================


def fetch_todo_progress_list(todo_id, user_guid=None):
    """
    获取 Todo 进度记录，用于判断进展描述是否相对上一次发生变化。
    data 兼容 list 或单条 dict。
    """
    payload = {
        "id": todo_id,
        "type": TODO_TYPE,
    }

    response = request_with_retry(
        "post",
        TODO_BASE_URL + TODO_PROGRESS_LIST_API,
        headers=get_todo_headers(user_guid=user_guid),
        json=payload,
        timeout=30,
    )

    response.raise_for_status()
    result = response.json()

    if result.get("code") not in (0, 200, "0", "200", None):
        print(f"⚠️ progressList 返回异常: todo={todo_id}, response={result}")
        return []

    data = result.get("data") or []

    if isinstance(data, list):
        return data

    if isinstance(data, dict):
        if "content" in data or "createBy" in data:
            return [data]
        for key in ("records", "list", "items"):
            if isinstance(data.get(key), list):
                return data.get(key)

    return []


# =============================================================================
# update todo
# =============================================================================


def update_todo(task):
    """
    更新 promiseDate / riskLevel / title。
    title 由主流程提前算好，避免 update_todo 内部重复判断。
    """
    payload = {
        "type": TODO_TYPE,
        "id": task["id"],
    }

    if task.get("promiseDate") is not None:
        payload["promiseDate"] = task["promiseDate"]

    if task.get("riskLevel") is not None:
        payload["riskLevel"] = task["riskLevel"]

    if task.get("title") is not None:
        payload["title"] = task["title"]

    if "promiseDate" not in payload and "riskLevel" not in payload and "title" not in payload:
        print(f"ℹ️ 无字段需更新，跳过 update: id={task['id']}")
        return None

    owner_user_guid = task.get("owner_user_guid")
    owner_display_name = task.get("owner_display_name", "")

    print(
        f"\n🚀 更新 Todo 字段: id={task['id']}, "
        f"owner={owner_display_name}, owner_user_guid={owner_user_guid}"
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))

    response = request_with_retry(
        "post",
        TODO_BASE_URL + TODO_UPDATE_API,
        headers=get_todo_headers(user_guid=owner_user_guid),
        json=payload,
        timeout=30,
    )

    response.raise_for_status()
    result = response.json()

    if result.get("code") not in (0, 200, "0", "200", None):
        raise Exception(f"Todo 更新接口返回异常: response={result}")

    print(f"✅ 更新成功: todo={task['id']}")
    return result


def insert_todo_progress(task):
    payload = {
        "id": task["id"],
        "type": TODO_TYPE,
        "content": task["progress_content"],
    }

    progress_user_guid = task.get("progress_user_guid")

    print(f"\n🚀 新增进展: id={task['id']}, progress_user_guid={progress_user_guid}")
    print(json.dumps(payload, ensure_ascii=False, indent=2))

    response = request_with_retry(
        "post",
        TODO_BASE_URL + TODO_INSERT_PROGRESS_API,
        headers=get_todo_headers(user_guid=progress_user_guid),
        json=payload,
        timeout=30,
    )

    response.raise_for_status()
    result = response.json()

    if result.get("code") not in (0, 200, "0", "200", None):
        raise Exception(f"新增进展接口返回异常: response={result}")

    print(f"✅ 新增进展成功: todo={task['id']}")
    return result

# =============================================================================
# 主流程
# =============================================================================

print("\n==========================================")
print("🚀 Todo 日报回写应用启动 (v5)")
print("==========================================\n")

print(f"BASE_URL: {BASE_URL}")
print(f"TODO_BASE_URL: {TODO_BASE_URL}")
print(f"TARGET_DATE: {TARGET_DATE}")
print(f"USER_GUID: {USER_GUID}")
print(f"MAPPING_GUID: {MAPPING_GUID}")
print(f"TODO_PROJECT_ID: {TODO_PROJECT_ID}")

mapping_rows = load_mapping_from_guid(MAPPING_GUID)

todo_list = fetch_todo_list()
if not todo_list:
    print("ℹ️ 当前未获取到 Todo 数据，流程结束")
    sys.exit(0)

todo_index = build_todo_index(todo_list)
dept_user_guid_map = build_dept_user_guid_map(todo_list)

visited_notes = set()

total_projects = 0
hit_notes = 0
total_parsed_tasks = 0
total_update_tasks = 0
progress_success_count = 0
progress_fail_count = 0
update_success_count = 0
update_fail_count = 0
skip_default_total = 0
skip_no_match_count = 0
skip_no_change_count = 0
card_success_count = 0
card_fail_count = 0
card_skip_count = 0

all_todo_ids = set(todo_index.keys())
collected_note_todo_ids = set()

success_update_records = []
unupdated_records = []
format_error_records = []
writeback_failed_records = []
update_fail_records = []


def _get_owner_dept(task):
    src_todo = todo_index.get(task.get("id"))
    if src_todo:
        for o in src_todo.get("owners", []) or []:
            if o.get("userGuid") == task.get("progress_user_guid"):
                return o.get("fullDeptName", "")
    return ""


for row in mapping_rows:
    dept = row.get("dept", "").strip()
    project_guid = row.get("project_guid")
    folder_guid = row.get("work_log_folder_guid")
    project_name = row.get("project_name", "")

    if not project_guid:
        continue

    if not folder_guid:
        continue

    total_projects += 1

    print(f"\n🚀 处理项目: {project_name}")

    candidate_user_guids = get_project_access_user_guids(dept, dept_user_guid_map)

    if not candidate_user_guids:
        print(f"⚠️ 项目缺少可访问身份，跳过: {project_name}")
        continue

    access_result = try_get_daily_note_and_doc(
        project_guid=project_guid,
        folder_guid=folder_guid,
        target_date=TARGET_DATE,
        project_name=project_name,
        candidate_user_guids=candidate_user_guids,
    )

    if not access_result:
        print(f"⚠️ 所有候选用户均无法访问项目日报: {project_name}")
        continue

    daily_note = access_result["daily_note"]
    doc_json = access_result["doc_json"]

    note_guid = daily_note["note_guid"]

    if note_guid in visited_notes:
        print(f"ℹ️ 日报已处理过，跳过重复 note: {note_guid}")
        continue

    visited_notes.add(note_guid)
    hit_notes += 1

    print(f"✅ 命中日报: {daily_note['note_title']}")

    parsed_tasks, skip_default_count = parse_tasks_from_doc(doc_json)
    skip_default_total += skip_default_count

    if not parsed_tasks:
        print("ℹ️ 当前日报无有效 Todo 更新")
        continue

    for parsed_task in parsed_tasks:
        if parsed_task.get("id") is not None:
            collected_note_todo_ids.add(parsed_task.get("id"))

    total_parsed_tasks += len(parsed_tasks)

    print("\n======================")
    print("日报解析结果")
    print("======================")
    print(json.dumps(parsed_tasks, ensure_ascii=False, indent=2))

    tasks_by_todo_id = OrderedDict()
    for parsed_task in parsed_tasks:
        tid = parsed_task.get("id")
        if tid not in tasks_by_todo_id:
            tasks_by_todo_id[tid] = []
        tasks_by_todo_id[tid].append(parsed_task)

    for tid, group_tasks in tasks_by_todo_id.items():
        source_todo = todo_index.get(tid)

        enriched_tasks = []
        for parsed_task in group_tasks:
            enriched_task = enrich_task_with_todo_context(
                parsed_task,
                todo_index,
                note_project_name=project_name,
            )

            if not enriched_task:
                if parsed_task.get("id") not in todo_index:
                    skip_no_match_count += 1
                else:
                    skip_no_change_count += 1
                continue

            enriched_tasks.append(enriched_task)

        if not enriched_tasks:
            continue

        total_update_tasks += len(enriched_tasks)

        print("\n======================")
        print(f"待处理 Todo (id={tid}): {len(enriched_tasks)} 条")
        print("======================")
        print(json.dumps(enriched_tasks, ensure_ascii=False, indent=2))

        # ------------------------------------------------------------
        # A. 报告统计 + 进展回写 + 格式错误卡片：按每个 owner-task
        # ------------------------------------------------------------
        for task in enriched_tasks:
            status_info = classify_owner_task_status(task)

            owner_name = task.get("progress_owner_display") or task.get("owner_display_name") or ""
            owner_dept = _get_owner_dept(task)
            real_title = strip_title_prefixes(task.get("source_todo_title", ""))

            # 成功更新不再按单个 owner 三项完整直接判断。
            # v6 规则：同一任务多 owner 合并后，三项均有填写且至少一项相对上次变化，才视为更新。
            # success_update_records 会在任务级 group_update_status 中统一追加。

            # 报告未更新详情在拿到 progressList 后统一生成，
            # 这里不再提前 append，避免无法判断“和上次填写内容一致”。

            # 格式错误不再单独成表，统一进入“未更新详情”的原因说明。

            if task.get("format_errors"):
                sent = send_format_error_card(task)
                if sent:
                    card_success_count += 1
                else:
                    card_fail_count += 1
            else:
                card_skip_count += 1

            # 进展回写不能在这里执行。
            # 需要先读取 progressList 并完成任务级有效更新判定，
            # 否则刚插入的今日进展会被 progressList 当成“最新历史记录”，
            # 导致 progress_changed 被误判为 False。

        # ------------------------------------------------------------
        # A2. 纯任务级汇总：
        #     不再把缺席 owner 作为单独未更新记录 append 到报告。
        #     任务是否未更新，只在后面的 group_is_updated 判定后追加一行。
        # ------------------------------------------------------------
        if source_todo:
            missing_owner_records = collect_missing_owner_unupdated_records(
                source_todo,
                group_tasks,
            )
            if missing_owner_records:
                print(
                    f"ℹ️ 多 owner 缺席仅参与任务级原因判断，不单独入表: todo={tid}, "
                    f"missing_owner_count={len(missing_owner_records)}"
                )

        # ------------------------------------------------------------
        # B. 字段更新：
        #    promiseDate/riskLevel 按 owner 顺序顺延；
        #    title 只看第一 owner。
        # ------------------------------------------------------------
        if not source_todo:
            continue

        owners = source_todo.get("owners", []) or []

        final_promise = None
        promise_contributor = None

        final_risk = None
        risk_contributor = None

        for owner in owners:
            owner_name_val = str(owner.get("name") or "").strip()
            owner_english_name = str(owner.get("englishName") or "").strip()

            matched_pt = None

            for pt in group_tasks:
                pt_name = str(pt.get("owner_name") or "").strip()

                if pt_name and (pt_name == owner_name_val or pt_name == owner_english_name):
                    matched_pt = pt
                    break

            if matched_pt:
                if final_promise is None and matched_pt.get("promiseDate") is not None:
                    final_promise = matched_pt.get("promiseDate")
                    promise_contributor = owner

                if final_risk is None and matched_pt.get("riskLevel") is not None:
                    final_risk = matched_pt.get("riskLevel")
                    risk_contributor = owner

            if final_promise is not None and final_risk is not None:
                break

        current_promise = normalize_date_for_compare(source_todo.get("promiseDate"))
        current_risk_val = str(source_todo.get("riskLevel") or "").strip()

        candidate_final_promise = final_promise
        candidate_final_risk = final_risk

        promise_changed = (
            candidate_final_promise is not None
            and normalize_date_for_compare(candidate_final_promise) != current_promise
        )
        risk_changed = (
            candidate_final_risk is not None
            and candidate_final_risk != current_risk_val
        )

        # 用 progressList 判断“用户实际进展”是否相对上一次发生变化。
        # 注意：只比较用户实际写入的进展，不比较字段一致提醒/格式错误提醒/状态变更提醒。
        try:
            progress_records = fetch_todo_progress_list(
                tid,
                user_guid=get_first_owner_user_guid(source_todo),
            )
        except Exception as e:
            print(f"⚠️ progressList 获取失败，进展变化判断降级为有填写即变化: todo={tid}, error={e}")
            progress_records = []

        latest_progress_by_owner = get_latest_progress_by_owner(progress_records)

        # v8：报告改为纯任务级汇总。
        # 不再在 owner-task 层提前 append 未更新记录；
        # 等任务级 group_is_updated 判定完成后，只对“未形成有效更新”的任务 append 一行。
        progress_changed = False
        has_group_progress = False
        progress_contributor_tasks = []

        for et in enriched_tasks:
            actual_progress = et.get("actual_progress", "")
            if not is_meaningful_progress(actual_progress):
                continue

            has_group_progress = True
            progress_contributor_tasks.append(et)

            previous_progress = latest_progress_by_owner.get(et.get("progress_user_guid"), "")

            if progress_records:
                if has_progress_changed(actual_progress, previous_progress):
                    progress_changed = True
            else:
                # progressList 失败或无历史记录时，只要有实际进展，视为变化。
                progress_changed = True

        has_group_promise = candidate_final_promise is not None
        has_group_risk = candidate_final_risk is not None

        has_group_format_error = any(
            bool(et.get("promiseInvalid") or et.get("riskInvalid"))
            for et in enriched_tasks
        )

        group_complete = (
            has_group_progress
            and has_group_promise
            and has_group_risk
            and not has_group_format_error
        )
        group_changed = progress_changed or promise_changed or risk_changed
        group_is_updated = group_complete and group_changed

        print(
            f"🧭 更新判定 todo={tid}: "
            f"complete={group_complete}, changed={group_changed}, "
            f"progress_changed={progress_changed}, promise_changed={promise_changed}, risk_changed={risk_changed}"
        )

        if not group_is_updated:
            task_reason = build_task_level_unupdated_reason(
                has_group_progress=has_group_progress,
                has_group_promise=has_group_promise,
                has_group_risk=has_group_risk,
                has_group_format_error=has_group_format_error,
                group_changed=group_changed,
                progress_changed=progress_changed,
                promise_changed=promise_changed,
                risk_changed=risk_changed,
                enriched_tasks=enriched_tasks,
                latest_progress_by_owner=latest_progress_by_owner,
                current_promise=current_promise,
                current_risk=current_risk_val,
            )

            unupdated_records.append({
                "todo_id": tid,
                "owner_name": "、".join(collect_owner_display_from_todo(source_todo)),
                "owner_dept": "",
                "todo_code": get_todo_code_for_display(source_todo),
                "task_title": strip_title_prefixes(source_todo.get("title", "")),
                "missing_fields": task_reason,
                "days_not_updated": calc_days_since_modify(source_todo.get("lastModifyDate", "")),
            })

        # ------------------------------------------------------------
        # A3. 进展回写：
        #     必须放在 progressList 历史读取和 group_is_updated 判定之后。
        #     否则本次 insert 的进展会污染“上一条进展”的比较基准。
        # ------------------------------------------------------------
        for task in enriched_tasks:
            if not task.get("progress_content"):
                continue

            owner_name = task.get("progress_owner_display") or task.get("owner_display_name") or ""
            owner_dept = _get_owner_dept(task)
            real_title = strip_title_prefixes(task.get("source_todo_title", ""))

            try:
                insert_todo_progress(task)
                progress_success_count += 1
            except Exception as e:
                progress_fail_count += 1
                print(f"❌ 新增进展失败 todo={task['id']} error={e}")
                print(traceback.format_exc())
                update_fail_records.append({
                    "owner_name": owner_name,
                    "owner_dept": owner_dept,
                    "todo_code": task.get("todo_code", ""),
                    "task_title": real_title,
                    "reason": f"新增进展失败: {e}",
                })

        # 与当前值一致则不传字段，但上面的 changed 判断已经保留。
        if final_promise is not None and normalize_date_for_compare(final_promise) == current_promise:
            final_promise = None
            promise_contributor = None

        if final_risk is not None and final_risk == current_risk_val:
            final_risk = None
            risk_contributor = None

        # title 按任务级合并判定：
        # 多 owner 合起来满足“完整 + 至少一项变化”就视为更新，并清除标签；
        # 否则加【X次未更新】。格式错误则额外加【格式错误】。
        group_status_for_title = {
            "is_success_update": group_is_updated,
            "is_unupdated": not group_is_updated,
            "is_format_error": has_group_format_error,
            "missing_fields": [],
            "error_fields": [],
        }

        new_title = build_update_title_v2(
            source_title=source_todo.get("title", ""),
            last_modify_date=source_todo.get("lastModifyDate", ""),
            first_owner_status=group_status_for_title,
            is_writeback_failed=False,
        )

        # 报告不再展示成功更新明细，group_is_updated 仅用于 title 标签判断。
        # 字段回写和进展回写逻辑保持不变。

        owner_fields_map = OrderedDict()

        if final_promise is not None and promise_contributor:
            pg = str(promise_contributor.get("userGuid") or "").strip()
            if pg:
                owner_fields_map.setdefault(pg, {
                    "owner_obj": promise_contributor,
                    "promiseDate": None,
                    "riskLevel": None,
                    "title": None,
                })
                owner_fields_map[pg]["promiseDate"] = final_promise

        if final_risk is not None and risk_contributor:
            rg = str(risk_contributor.get("userGuid") or "").strip()
            if rg:
                owner_fields_map.setdefault(rg, {
                    "owner_obj": risk_contributor,
                    "promiseDate": None,
                    "riskLevel": None,
                    "title": None,
                })
                owner_fields_map[rg]["riskLevel"] = final_risk

        if owner_fields_map:
            first_update = True

            for user_guid, fields in owner_fields_map.items():
                owner_obj = fields["owner_obj"]
                owner_display = (
                    owner_obj.get("displayName")
                    or owner_obj.get("name")
                    or owner_obj.get("englishName")
                    or ""
                )
                owner_dept = str(owner_obj.get("fullDeptName") or "")

                update_task = {
                    "id": tid,
                    "type": TODO_TYPE,
                    "promiseDate": fields.get("promiseDate"),
                    "riskLevel": fields.get("riskLevel"),
                    "title": new_title if first_update else None,
                    "owner_user_guid": user_guid,
                    "owner_display_name": owner_display,
                }

                try:
                    result = update_todo(update_task)
                    if result is not None:
                        update_success_count += 1
                except Exception as e:
                    update_fail_count += 1
                    print(f"❌ 字段/title 更新失败 todo={tid} error={e}")
                    print(traceback.format_exc())
                    update_fail_records.append({
                        "owner_name": owner_display,
                        "owner_dept": owner_dept,
                        "todo_code": get_todo_code_for_display(source_todo),
                        "task_title": strip_title_prefixes(source_todo.get("title", "")),
                        "reason": f"字段/title 更新失败: {e}",
                    })

                first_update = False

        elif new_title is not None:
            first_owner = get_first_owner_identity(source_todo)
            if first_owner and first_owner.get("user_guid"):
                try:
                    result = update_todo({
                        "id": tid,
                        "type": TODO_TYPE,
                        "promiseDate": None,
                        "riskLevel": None,
                        "title": new_title,
                        "owner_user_guid": first_owner["user_guid"],
                        "owner_display_name": first_owner["display_name"],
                    })
                    if result is not None:
                        update_success_count += 1
                except Exception as e:
                    update_fail_count += 1
                    print(f"❌ title 更新失败 todo={tid} error={e}")
                    print(traceback.format_exc())
                    update_fail_records.append({
                        "owner_name": first_owner.get("display_name", ""),
                        "owner_dept": first_owner.get("dept", ""),
                        "todo_code": get_todo_code_for_display(source_todo),
                        "task_title": strip_title_prefixes(source_todo.get("title", "")),
                        "reason": f"title 更新失败: {e}",
                    })
            else:
                print(f"⚠️ 无法更新 title，Todo 缺少第一 owner userGuid: todo={tid}")

# =============================================================================
# C. 失败任务处理：TodoList 有，但日报未找到 todo_id
# =============================================================================

missing_todo_ids = all_todo_ids - collected_note_todo_ids

print(f"\n🔎 回写失败任务检查: TodoList总数={len(all_todo_ids)}, 日报收集到={len(collected_note_todo_ids)}, 缺失={len(missing_todo_ids)}")

for missing_id in sorted(missing_todo_ids):
    source_todo = todo_index.get(missing_id)
    if not source_todo:
        continue

    real_title = strip_title_prefixes(source_todo.get("title", ""))
    todo_code = get_todo_code_for_display(source_todo)

    owners = source_todo.get("owners", []) or []

    if not owners:
        writeback_failed_records.append({
            "owner_name": "",
            "owner_dept": "",
            "todo_code": todo_code,
            "task_title": real_title,
            "todo_id": missing_id,
            "reason": "TodoList中存在该任务，但日报未找到对应任务ID，且任务无 owner",
        })
    else:
        for owner in owners:
            writeback_failed_records.append({
                "owner_name": owner.get("name") or owner.get("englishName") or owner.get("displayName") or "",
                "owner_dept": owner.get("fullDeptName", ""),
                "todo_code": todo_code,
                "task_title": real_title,
                "todo_id": missing_id,
                "reason": "TodoList中存在该任务，但日报未找到对应任务ID",
            })

    new_title = build_update_title_v2(
        source_title=source_todo.get("title", ""),
        last_modify_date=source_todo.get("lastModifyDate", ""),
        first_owner_status={
            "is_success_update": False,
            "is_unupdated": False,
            "is_format_error": False,
            "missing_fields": [],
            "error_fields": [],
        },
        is_writeback_failed=True,
    )

    first_owner = get_first_owner_identity(source_todo)

    if new_title is not None and first_owner and first_owner.get("user_guid"):
        try:
            result = update_todo({
                "id": missing_id,
                "type": TODO_TYPE,
                "promiseDate": None,
                "riskLevel": None,
                "title": new_title,
                "owner_user_guid": first_owner["user_guid"],
                "owner_display_name": first_owner["display_name"],
            })
            if result is not None:
                update_success_count += 1
        except Exception as e:
            update_fail_count += 1
            print(f"❌ 回写失败标签更新失败 todo={missing_id}, error={e}")
            print(traceback.format_exc())
            update_fail_records.append({
                "owner_name": first_owner.get("display_name", ""),
                "owner_dept": first_owner.get("dept", ""),
                "todo_code": todo_code,
                "task_title": real_title,
                "reason": f"回写失败标签更新失败: {e}",
            })

# 真实 update 接口失败也并入“更新失败人员情况”
writeback_failed_records.extend(update_fail_records)

print("\n==========================================")
print("🏁 Todo 日报回写应用完成")
print("==========================================\n")

print("================ 执行统计 ================")
print(f"处理项目数: {total_projects}")
print(f"命中日报数: {hit_notes}")
print(f"TodoList任务总数: {len(todo_index)}")
print(f"TodoList owner总数: {count_unique_todo_owners(todo_list)}")
print(f"解析任务数: {total_parsed_tasks}")
print(f"待处理 owner-task 数: {total_update_tasks}")
print(f"成功更新明细已取消展示，任务级更新仅用于 title 判断")
print(f"未形成有效更新任务: {len(unupdated_records)}")
print(f"格式错误已并入任务级未更新说明: {len(format_error_records)}")
print(f"失败/回写失败记录: {len(writeback_failed_records)}")
print(f"新增进展成功: {progress_success_count}")
print(f"新增进展失败: {progress_fail_count}")
print(f"字段/title更新成功: {update_success_count}")
print(f"字段/title更新失败: {update_fail_count}")
print(f"跳过(默认模板): {skip_default_total}")
print(f"跳过(未匹配 Todo): {skip_no_match_count}")
print(f"跳过(无有效变化/缺少owner): {skip_no_change_count}")
print(f"格式错误卡片发送成功: {card_success_count}")
print(f"格式错误卡片发送失败/跳过: {card_fail_count}")
print(f"无格式错误无需发卡片: {card_skip_count}")
print("==========================================\n")

# =============================================================================
# D. 报告记录最终去重与 title 清洗
# =============================================================================

success_update_records = dedupe_records(
    success_update_records,
    lambda r: (
        r.get("todo_id"),
        r.get("owner_name"),
        r.get("owner_dept"),
    )
)

unupdated_records = dedupe_records(
    unupdated_records,
    lambda r: (
        r.get("todo_id"),
        r.get("missing_fields"),
    )
)

format_error_records = dedupe_records(
    format_error_records,
    lambda r: (
        r.get("todo_id"),
        r.get("owner_name"),
        r.get("owner_dept"),
        r.get("error_field"),
        r.get("error_raw"),
    )
)

writeback_failed_records = dedupe_records(
    writeback_failed_records,
    lambda r: (
        r.get("todo_id"),
        r.get("owner_name"),
        r.get("owner_dept"),
        r.get("reason"),
    )
)

for record_list in (
    success_update_records,
    unupdated_records,
    format_error_records,
    writeback_failed_records,
):
    for record in record_list:
        record["task_title"] = strip_title_prefixes(record.get("task_title", ""))


report_data = {
    "total_todo_tasks": len(todo_index),
    "total_todo_owners": count_unique_todo_owners(todo_list),
    "success_update_records": success_update_records,
    "unupdated_records": unupdated_records,
    "format_error_records": format_error_records,
    "writeback_failed_records": writeback_failed_records,
}

report_result = write_report(TARGET_DATE, report_data)
if report_result:
    doc_guid, report_title = report_result
    send_report_message(report_title, doc_guid, report_data)

if progress_fail_count > 0 or update_fail_count > 0:
    raise Exception(
        f"存在 {progress_fail_count} 个新增进展失败、"
        f"{update_fail_count} 个字段/title更新失败，请查看日志"
    )
